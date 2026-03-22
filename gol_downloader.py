"""
Garden of Life Image Downloader — Using Crawlee + Playwright
=============================================================
Uses a real headless Chrome browser to bypass 403 bot protection.
Downloads images 1, 2, 3 for each product using UPC.

Usage:
    python gol_downloader.py --input "Product_Image_Resize_.xlsx"

Install:
    pip install "crawlee[playwright]" openpyxl pillow numpy rembg onnxruntime
    playwright install chromium
"""

import os, sys, re, asyncio, argparse
from pathlib import Path
from io import BytesIO
from openpyxl import load_workbook
from PIL import Image, ImageFilter, ImageEnhance
import numpy as np

# Save images in same folder as the script
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resized_images")
GOL_DOMAIN  = "gardenoflife.com"

SIZES = {
    "500x500_MasterDB":   (500,  500),
    "1000x1000_LiveSite": (1000, 1000),
}


# ── Background removal ────────────────────────────────────────────────────────
def has_clean_background(img, threshold=0.50):
    arr   = np.array(img.convert("RGB"))
    white = (arr[:,:,0] > 235) & (arr[:,:,1] > 235) & (arr[:,:,2] > 235)
    return white.sum() / white.size > threshold


def remove_background(img):
    if has_clean_background(img):
        print(f"    [BG] Clean background — skipping rembg")
        return img.convert("RGBA")
    print(f"    [BG] Removing background...")
    try:
        from rembg import remove as rembg_remove, new_session
        session = new_session("u2net")
        buf = BytesIO()
        img.save(buf, format="PNG")
        result = rembg_remove(buf.getvalue(), session=session, alpha_matting=False)
        return Image.open(BytesIO(result)).convert("RGBA")
    except Exception as e:
        print(f"    [BG] Failed ({e}) — using original")
        return img.convert("RGBA")


# ── Process and save image ────────────────────────────────────────────────────
def process_image_bytes(img_bytes, folder, name):
    """Process raw image bytes and save to 4 sizes."""
    try:
        img = Image.open(BytesIO(img_bytes))
        if img.width < 100 or img.height < 100:
            print(f"    Too small ({img.width}x{img.height}) — skipping")
            return False

        print(f"    Downloaded: {img.width}x{img.height}px")
        img = remove_background(img)

        # Tight crop using alpha
        arr  = np.array(img)
        mask = arr[:,:,3] > 10
        rows = np.any(mask, axis=1)
        cols = np.any(mask, axis=0)
        if rows.any() and cols.any():
            rmin = int(np.where(rows)[0][0]);  rmax = int(np.where(rows)[0][-1])
            cmin = int(np.where(cols)[0][0]);  cmax = int(np.where(cols)[0][-1])
            ph = max(4, int((rmax - rmin) * 0.02))
            pw = max(4, int((cmax - cmin) * 0.02))
            img = img.crop((max(0,cmin-pw), max(0,rmin-ph),
                            min(img.width,cmax+pw+1), min(img.height,rmax+ph+1)))
            print(f"    [Crop] {img.width}x{img.height}")

        os.makedirs(folder, exist_ok=True)
        for label, (w, h) in SIZES.items():
            scale   = min(int(w*0.90)/img.width, int(h*0.90)/img.height)
            new_w   = int(img.width  * scale)
            new_h   = int(img.height * scale)
            resized = img.resize((new_w, new_h), Image.LANCZOS)

            rgb = resized.convert("RGB")
            rgb = rgb.filter(ImageFilter.UnsharpMask(radius=1.0, percent=120, threshold=3))
            rgb = ImageEnhance.Sharpness(rgb).enhance(1.3)
            r2,g2,b2 = rgb.split(); _,_,_,a = resized.split()
            resized = Image.merge("RGBA",(r2,g2,b2,a))

            canvas = Image.new("RGBA",(w,h),(255,255,255,255))
            canvas.paste(resized,((w-new_w)//2,(h-new_h)//2),resized)
            out = os.path.join(folder, f"{name}_{label}.jpg")
            canvas.convert("RGB").save(out, "JPEG", quality=95, dpi=(96,96))
            print(f"      → {os.path.basename(out)}")

        return True
    except Exception as e:
        print(f"    Error processing image: {e}")
        return False


# ── Garden of Life downloader using Crawlee ───────────────────────────────────
async def download_gol_products(products):
    """
    products: list of (sku, upc, folder) tuples
    Uses Crawlee PlaywrightCrawler — real headless Chrome, bypasses 403.
    """
    from crawlee.crawlers import PlaywrightCrawler, PlaywrightCrawlingContext

    # Build list of all image URLs to download
    # Each item: (url, folder, filename_stem)
    tasks = []
    for sku, upc, folder in products:
        upc = str(upc).strip()
        base = f"https://www.gardenoflife.com/media/catalog/product/{upc[0]}/{upc[1]}"
        for n in range(1, 4):
            url   = f"{base}/{upc}-{n}_1.jpg?optimize=medium&fit=bounds&height=2000&width=2000"
            fname = f"{upc}-{n}_1"
            tasks.append((url, folder, fname, sku, upc, n))

    # Results dict: url -> bytes or None
    results = {}

    crawler = PlaywrightCrawler(
        headless=True,
        browser_type="chromium",
        max_requests_per_crawl=len(tasks) + 1,
    )

    @crawler.router.default_handler
    async def handler(context: PlaywrightCrawlingContext) -> None:
        url = context.request.url
        try:
            # Use page.goto to load the image URL directly
            response = await context.page.goto(url)
            if response and response.status == 200:
                body = await response.body()
                results[url] = body
                print(f"  ✓ Got {len(body)} bytes: {url.split('/')[-1].split('?')[0]}")
            elif response:
                print(f"  HTTP {response.status}: {url.split('/')[-1].split('?')[0]}")
                results[url] = None
            else:
                results[url] = None
        except Exception as e:
            print(f"  Error fetching {url}: {e}")
            results[url] = None

    # Run crawler with all image URLs
    urls = [t[0] for t in tasks]
    print(f"\n  [GOL] Launching headless Chrome for {len(urls)} images...")
    await crawler.run(urls)

    # Process downloaded images
    downloaded_count = {}  # sku -> count
    for url, folder, fname, sku, upc, n in tasks:
        img_bytes = results.get(url)
        if img_bytes and len(img_bytes) > 1000:
            print(f"\n  [{n}/3] Processing {fname}.jpg for SKU {sku}...")
            if process_image_bytes(img_bytes, folder, fname):
                downloaded_count[sku] = downloaded_count.get(sku, 0) + 1
        else:
            if n == 1:
                print(f"  [{n}/3] {fname}.jpg — not available")

    return downloaded_count


# ── Excel processing ──────────────────────────────────────────────────────────
def process_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb["Image Resize Jobs"]

    gol_products = []   # Garden of Life rows
    row_map      = {}   # sku -> row for status update

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        upc    = str(row[1].value).strip() if row[1].value else ""
        sku    = str(row[2].value).strip() if row[2].value else ""
        url    = str(row[5].value).strip() if row[5].value else ""
        status = str(row[6].value).strip() if row[6].value else ""

        if not upc or upc == "UPC *": continue
        if status == "Done":
            print(f"Skipping {sku or upc} — already Done"); continue
        if not url: continue
        if GOL_DOMAIN not in url: continue
        if not upc:
            print(f"Skipping SKU {sku} — no UPC"); continue

        identifier = sku if sku else upc
        folder = os.path.join(OUTPUT_DIR, identifier)
        os.makedirs(folder, exist_ok=True)

        gol_products.append((sku, upc, folder))
        row_map[sku or upc] = row[6]  # status cell

    if not gol_products:
        print("No Garden of Life products found in Excel.")
        return

    print(f"\nFound {len(gol_products)} Garden of Life products to process")

    # Run async downloader
    downloaded = asyncio.run(download_gol_products(gol_products))

    # Update Excel status
    for sku, upc, folder in gol_products:
        key    = sku or upc
        count  = downloaded.get(sku, 0)
        cell   = row_map.get(key)
        if cell:
            cell.value = "Done" if count > 0 else "Failed"
        print(f"  SKU {sku}: {count} images downloaded — {'Done' if count > 0 else 'Failed'}")

    wb.save(filepath)
    print(f"\nComplete! Results saved to Excel.")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Garden of Life Image Downloader")
    parser.add_argument("--input", help="Path to Excel file")
    parser.add_argument("--upc",   help="Single product UPC")
    parser.add_argument("--name",   help="Product name (ignored by this plugin)")
    parser.add_argument("--sku",   help="Single product SKU")
    parser.add_argument("--folder",help="Output folder for single product")
    args = parser.parse_args()

    if args.upc and args.folder:
        # Single product mode — called from download_webp.py plugin
        sku    = args.sku or args.upc
        folder = args.folder
        os.makedirs(folder, exist_ok=True)
        products = [(sku, args.upc, folder)]
        downloaded = asyncio.run(download_gol_products(products))
        count = downloaded.get(sku, 0)
        print(f"  SKU {sku}: {count} images downloaded")
        sys.exit(0 if count > 0 else 1)

    elif args.input:
        if not os.path.exists(args.input):
            print(f"File not found: {args.input}"); sys.exit(1)
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        process_excel(args.input)

    else:
        print("Usage: gol_downloader.py --input file.xlsx")
        print("   or: gol_downloader.py --upc UPC --sku SKU --folder FOLDER")
        sys.exit(1)
