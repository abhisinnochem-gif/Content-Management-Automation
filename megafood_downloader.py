"""
MegaFood (megafood.com) Image Downloader
=========================================
Shopify store — pages through /collections/all/products.json to find
product by SKU. Uses variant.featured_image.src for exact bottle image.

Image selection:
  - Main : variant.featured_image.src — exact image for this SKU/size
  - SFP  : image with empty variant_ids[] whose filename contains
           'supplement' and 'fact' — cropped to remove:
             * bottom green footer bar (NON GMO / badges row)
             * right-side green banner column (badge column)

Usage:
    python megafood_downloader.py --sku 10006 --upc 051494100066 --name "Turmeric Strength" --folder resized_images/10006
    python megafood_downloader.py --input "Product_Image_Resize_.xlsx"
"""

import os, sys, argparse, requests, re
from io import BytesIO
from openpyxl import load_workbook
from PIL import Image, ImageFilter, ImageEnhance
import numpy as np

OUTPUT_DIR      = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resized_images")
MEGAFOOD_DOMAIN = "megafood.com"
BASE_URL        = "https://megafood.com"

SIZES = {
    "500x500_MasterDB":   (500,  500),
    "1000x1000_LiveSite": (1000, 1000),
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Accept": "application/json",
}


# ── Supplement facts crop ─────────────────────────────────────────────────────
def crop_supplement_facts(img):
    """
    Remove dark green regions from supplement facts image:
      1. Right-side green banner column (badge column)
      2. Bottom green footer bar (NON GMO / Vegetarian etc.)
    Dark MegaFood green: R<100, G>80, B<100
    """
    arr = np.array(img.convert("RGB"))
    h, w = arr.shape[:2]

    def is_green_dominant(pixels):
        """Returns ratio of dark-green pixels in a 1D array of RGB pixels."""
        return np.sum((pixels[:,1] > 80) & (pixels[:,0] < 100) & (pixels[:,2] < 100)) / len(pixels)

    # ── 1. Crop right-side green column ──────────────────────────────────────
    # Only crop if green band is at least 5% of image width (real banner, not border)
    green_start_x = w
    for x in range(w - 1, int(w * 0.50), -1):
        col = arr[:, x, :]
        if is_green_dominant(col) > 0.30:
            green_start_x = x
        else:
            break  # stop at first non-green column from right

    green_band_width = w - green_start_x
    if green_band_width > w * 0.05:
        print(f"    [Crop SF] Removing right green column x={green_start_x} (band={green_band_width}px)")
        img = img.crop((0, 0, green_start_x, h))
        arr = np.array(img.convert("RGB"))
        h, w = arr.shape[:2]
    else:
        print(f"    [Crop SF] No right green column detected (band={green_band_width}px < 5%)")

    # ── 2. Crop bottom green footer bar ──────────────────────────────────────
    crop_y = h
    for y in range(h - 1, int(h * 0.50), -1):
        row = arr[y]
        if is_green_dominant(row) > 0.30:
            crop_y = y
        elif crop_y < h:
            break

    if crop_y < h:
        print(f"    [Crop SF] Removing bottom green footer at y={crop_y} of {h}")
        img = img.crop((0, 0, w, crop_y))
    else:
        print(f"    [Crop SF] No bottom green footer detected")

    return img


# ── Background removal ────────────────────────────────────────────────────────
def has_clean_background(img, threshold=0.50):
    arr   = np.array(img.convert("RGB"))
    # Check for white OR light grey background (MegaFood uses both)
    # White: all channels > 235
    # Light grey: all channels > 200 and channels are close to each other (neutral)
    white     = (arr[:,:,0] > 235) & (arr[:,:,1] > 235) & (arr[:,:,2] > 235)
    light_grey = (arr[:,:,0] > 200) & (arr[:,:,1] > 200) & (arr[:,:,2] > 200) &                  (np.abs(arr[:,:,0].astype(int) - arr[:,:,1].astype(int)) < 15) &                  (np.abs(arr[:,:,1].astype(int) - arr[:,:,2].astype(int)) < 15)
    clean = white | light_grey
    return clean.sum() / clean.size > threshold

def remove_background(img):
    if has_clean_background(img):
        print(f"    [BG] Clean background — skipping rembg")
        return img.convert("RGBA")
    print(f"    [BG] Removing background...")
    try:
        from rembg import remove as rembg_remove, new_session
        session = new_session("u2net")
        buf = BytesIO()
        img.convert("RGB").save(buf, format="PNG")
        result = rembg_remove(buf.getvalue(), session=session, alpha_matting=False)
        return Image.open(BytesIO(result)).convert("RGBA")
    except Exception as e:
        print(f"    [BG] Failed ({e}) — using original")
        return img.convert("RGBA")


# ── Process and save image ────────────────────────────────────────────────────
def process_image_bytes(img_bytes, folder, name, img_type="main"):
    try:
        img = Image.open(BytesIO(img_bytes))
        if img.width < 100 or img.height < 100:
            print(f"    Too small ({img.width}x{img.height}) — skipping")
            return False

        print(f"    Downloaded: {img.width}x{img.height}px")

        if img_type == "sfp":
            # Crop green regions only — keep full white panel, no rembg
            img = crop_supplement_facts(img)
            img = img.convert("RGBA")
            print(f"    [BG] No rembg — SFP white panel: {img.width}x{img.height}")
        else:
            # MegaFood images have clean white/grey backgrounds — no rembg needed
            # Just paste onto white canvas directly
            img = img.convert("RGBA")
            print(f"    [BG] No rembg — clean source image")

        os.makedirs(folder, exist_ok=True)

        for label, (w, h) in SIZES.items():
            scale   = min(int(w*0.90)/img.width, int(h*0.90)/img.height)
            new_w   = int(img.width  * scale)
            new_h   = int(img.height * scale)
            resized = img.resize((new_w, new_h), Image.LANCZOS)

            rgb = resized.convert("RGB")
            rgb = rgb.filter(ImageFilter.UnsharpMask(radius=1.0, percent=120, threshold=3))
            rgb = ImageEnhance.Sharpness(rgb).enhance(1.3)
            r2,g2,b2 = rgb.split(); _,_,_,a = resized.split()
            resized = Image.merge("RGBA",(r2,g2,b2,a))

            canvas = Image.new("RGBA",(w,h),(255,255,255,255))
            canvas.paste(resized,((w-new_w)//2,(h-new_h)//2),resized)
            out = os.path.join(folder, f"{name}_{label}.jpg")
            canvas.convert("RGB").save(out, "JPEG", quality=95, dpi=(96,96))
            print(f"      → {os.path.basename(out)}")

        return True
    except Exception as e:
        print(f"    Error processing image: {e}")
        return False


# ── Step 1: Page through products.json, match by SKU ─────────────────────────
def find_product_by_sku(sku, upc):
    """
    Pages through /collections/all/products.json.
    Matches variant by SKU field (exact). Returns (product, variant) or (None, None).
    """
    sku_upper = str(sku).strip().upper()
    upc_clean = str(upc).strip()
    page      = 1

    print(f"    Scanning products.json for SKU {sku}...")
    while True:
        url = f"{BASE_URL}/collections/all/products.json?limit=250&page={page}"
        try:
            r = requests.get(url, headers=HEADERS, timeout=20)
            if r.status_code != 200:
                print(f"    products.json HTTP {r.status_code} on page {page}")
                break

            products = r.json().get("products", [])
            if not products:
                print(f"    Scanned {page-1} page(s) — SKU not found")
                break

            print(f"    Page {page}: {len(products)} products")

            for product in products:
                for variant in product.get("variants", []):
                    v_sku     = str(variant.get("sku",     "")).strip().upper()
                    v_barcode = str(variant.get("barcode", "")).strip()
                    if v_sku == sku_upper or v_barcode == upc_clean:
                        print(f"    ✓ Found: '{product['title']}' (SKU={v_sku})")
                        return product, variant

            page += 1

        except Exception as e:
            print(f"    products.json error: {e}")
            break

    return None, None


# ── Step 2: Select main + SFP images ─────────────────────────────────────────
def get_images(product, variant):
    """
    Main : variant.featured_image.src — exact bottle for this SKU
    SFP  : image with empty variant_ids[] whose filename contains
           both 'supplement' and 'fact' (case-insensitive)
    """
    images   = product.get("images", [])
    found    = []
    main_src = None

    # ── Main: use variant.featured_image directly ─────────────────────────────
    fi = variant.get("featured_image")
    if fi and fi.get("src"):
        src      = fi["src"]
        fname    = src.split("/")[-1].split("?")[0]
        main_src = src.split("?")[0] + "?width=2000"
        found.append((main_src, fname, "main"))
        print(f"    Found [main] via featured_image: {fname}")
    else:
        # Fallback: position 1
        sorted_imgs = sorted(images, key=lambda x: x.get("position", 99))
        if sorted_imgs:
            src      = sorted_imgs[0]["src"]
            fname    = src.split("/")[-1].split("?")[0]
            main_src = src.split("?")[0] + "?width=2000"
            found.append((main_src, fname, "main"))
            print(f"    Found [main] via position 1 fallback: {fname}")

    # ── SFP: no variant_ids + filename contains 'supplement' + 'fact' ─────────
    for img in images:
        if img.get("variant_ids"):   # skip variant-specific images
            continue
        src      = img["src"]
        fname    = src.split("/")[-1].split("?")[0]
        fname_lw = fname.lower()
        if ("supplement" in fname_lw or "sup-fact" in fname_lw or "supfact" in fname_lw) and "fact" in fname_lw:
            sfp_src = src.split("?")[0] + "?width=2000"
            if sfp_src != main_src:
                found.append((sfp_src, fname, "sfp"))
                print(f"    Found [sfp]: {fname}")
                break

    return found


# ── Step 3: Download ──────────────────────────────────────────────────────────
def download_image(img_url):
    try:
        r = requests.get(img_url, headers=HEADERS, timeout=20)
        if r.status_code == 200 and len(r.content) > 1000:
            return r.content
        print(f"    HTTP {r.status_code}")
        return None
    except Exception as e:
        print(f"    Download error: {e}")
        return None


# ── Main: one product ─────────────────────────────────────────────────────────
def process_one(sku, upc, product_name, folder):
    product, variant = find_product_by_sku(sku, upc)
    if not product:
        print(f"  SKU {sku} — not found in MegaFood catalog")
        return 0

    images = get_images(product, variant)
    if not images:
        print(f"  SKU {sku} — no images found")
        return 0

    os.makedirs(folder, exist_ok=True)
    count = 0
    for img_url, fname, img_type in images:
        name      = f"{upc}_{img_type}"
        print(f"  Downloading [{img_type}] {fname}...")
        img_bytes = download_image(img_url)
        if img_bytes and process_image_bytes(img_bytes, folder, name, img_type):
            count += 1
    return count


# ── Excel processing ──────────────────────────────────────────────────────────
def process_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb["Image Resize Jobs"]
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        upc    = str(row[1].value).strip() if row[1].value else ""
        sku    = str(row[2].value).strip() if row[2].value else ""
        name   = str(row[3].value).strip() if row[3].value else ""
        url    = str(row[5].value).strip() if row[5].value else ""
        status = str(row[6].value).strip() if row[6].value else ""
        if not upc or upc == "UPC *": continue
        if status == "Done":
            print(f"Skipping {sku} — already Done"); continue
        if not url or MEGAFOOD_DOMAIN not in url: continue
        rows.append((sku, upc, name, row[6]))

    if not rows:
        print("No MegaFood products found."); return

    total = len(rows)
    print(f"\nFound {total} MegaFood products to process\n")
    for i, (sku, upc, name, status_cell) in enumerate(rows, 1):
        folder = os.path.join(OUTPUT_DIR, sku)
        os.makedirs(folder, exist_ok=True)
        print(f"{'='*60}")
        print(f"[{i}/{total}] SKU: {sku} | UPC: {upc} | {name}")
        print(f"{'='*60}")
        count = process_one(sku, upc, name, folder)
        status_cell.value = "Done" if count > 0 else "Failed"
        print(f"  → {count} images saved\n")
        wb.save(filepath)
    print(f"\nComplete!")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="MegaFood Image Downloader")
    parser.add_argument("--input",  help="Path to Excel file")
    parser.add_argument("--sku",    help="Single product SKU")
    parser.add_argument("--upc",    help="Single product UPC")
    parser.add_argument("--name",   help="Product name")
    parser.add_argument("--folder", help="Output folder")
    args = parser.parse_args()

    if args.sku and args.folder:
        os.makedirs(args.folder, exist_ok=True)
        upc   = args.upc  if args.upc  else args.sku
        name  = args.name if args.name else args.sku
        count = process_one(args.sku, upc, name, args.folder)
        print(f"  SKU {args.sku}: {count} images downloaded")
        sys.exit(0 if count > 0 else 1)
    elif args.input:
        if not os.path.exists(args.input):
            print(f"File not found: {args.input}"); sys.exit(1)
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        process_excel(args.input)
    else:
        print("Usage: megafood_downloader.py --input file.xlsx")
        print("   or: megafood_downloader.py --sku 10006 --upc 051494100066 --name 'Turmeric Strength' --folder resized_images/10006")
        sys.exit(1)
