"""
Auto Image Downloader
======================
Universal image downloader — uses sitemap as primary strategy (works on any
platform), falls back to platform-specific plugins if sitemap fails.

Strategy order:
  1. generic_sitemap.py  — sitemap crawl (universal, works on all platforms)
  2. Platform detection  — fallback only if sitemap finds nothing
     shopify     → generic_shopify.py
     woocommerce → generic_wordpress.py
     wordpress   → generic_wordpress.py
     magento     → generic_magento.py

Excel columns (row 3 = header, data starts row 4):
  B = UPC
  C = PTID        (used as output filename)
  D = Product Name
  E = Vendor Name
  F = Vendor Product URL
  G = Status      (auto-filled: Done / Failed)

Usage:
    python auto_download.py --input "Product_Image_Resize_.xlsx"
    python auto_download.py --url https://synevit.com --upc 860011318019 --ptid 223140 --name "Neurocomplex-B Slow-Release" --vendor "Synevit"
    python auto_download.py --input "Product_Image_Resize_.xlsx" --vendor "Synevit"
    python auto_download.py --url https://synevit.com --upc 860011318019 --ptid 223140 --name "Neurocomplex-B" --vendor "Synevit" --platform wordpress
    python auto_download.py --url https://synevit.com --upc 860011318019 --ptid 223140 --name "Neurocomplex-B" --vendor "Synevit" --product-url https://synevit.com/eu-products/neurocomplex-b/

Install:
    pip install requests beautifulsoup4 openpyxl pillow numpy rembg onnxruntime lxml
    pip install "crawlee[playwright]"
    playwright install chromium
"""

import os, sys, argparse
from urllib.parse import urlparse
from openpyxl import load_workbook

# ── Import plugins ────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from detect_platform import detect_platform
import generic_sitemap
import generic_shopify
import generic_wordpress
import generic_magento

OUTPUT_DIR  = os.path.join(SCRIPT_DIR, "resized_images")
EXCEL_SHEET = "Image Resize Jobs"

# Platform routing
SHOPIFY_PLATFORMS   = {"shopify"}
WORDPRESS_PLATFORMS = {"wordpress", "woocommerce", "weebly", "squarespace"}
MAGENTO_PLATFORMS   = {"magento"}


# ── Platform cache — avoid re-detecting same domain ──────────────────────────
_platform_cache = {}

def get_platform(url, force_platform=None):
    if force_platform:
        platform_map = {
            "shopify":     {"platform": "shopify",     "confidence": "forced", "signal": "CLI override"},
            "wordpress":   {"platform": "wordpress",   "confidence": "forced", "signal": "CLI override"},
            "woocommerce": {"platform": "woocommerce", "confidence": "forced", "signal": "CLI override"},
            "magento":     {"platform": "magento",     "confidence": "forced", "signal": "CLI override"},
        }
        key = force_platform.lower()
        if key in platform_map:
            print(f"  Platform: {key.upper()} (forced via --platform)")
            return platform_map[key]
    domain = urlparse(url).netloc.replace("www.", "")
    if domain not in _platform_cache:
        print(f"  Detecting platform for {domain}...")
        result = detect_platform(url)
        _platform_cache[domain] = result
        print(f"  Platform: {result['platform'].upper()} ({result['confidence']}) — {result['signal']}")
    return _platform_cache[domain]


# ── Route to correct plugin ───────────────────────────────────────────────────
def process_one(vendor_url, upc, product_name, vendor_name="Unknown", ptid=None,
                force_platform=None, **kwargs):
    """
    Download product image — sitemap first, platform plugins as fallback.

    Args:
      vendor_url     : vendor website URL
      upc            : product UPC barcode
      product_name   : product name
      vendor_name    : output subfolder name
      ptid           : output filename
      force_platform : skip sitemap, force a specific platform plugin
      product_url    : skip sitemap, go directly to this page
    """
    product_url = kwargs.get("product_url") or None

    # ── Forced platform — skip sitemap entirely ───────────────────────────────
    if force_platform:
        print(f"  Platform: {force_platform.upper()} (forced — skipping sitemap)")
        return _route_to_platform(force_platform, vendor_url, upc, product_name,
                                  vendor_name, ptid, product_url)

    # ── Strategy 1: Sitemap (universal — works on all platforms) ─────────────
    print(f"  Strategy: Sitemap (universal)")
    count = generic_sitemap.process_one(
        vendor_url   = vendor_url,
        upc          = upc,
        product_name = product_name,
        vendor_name  = vendor_name,
        ptid         = ptid,
        product_url  = product_url,
    )
    if count > 0:
        return count

    # ── Strategy 2: Platform-specific fallback ────────────────────────────────
    print(f"  Sitemap strategy failed — falling back to platform detection...")
    detection = get_platform(vendor_url)
    platform  = detection["platform"]
    return _route_to_platform(platform, vendor_url, upc, product_name,
                              vendor_name, ptid, product_url)


def _route_to_platform(platform, vendor_url, upc, product_name,
                       vendor_name, ptid, product_url=None):
    """Route to the correct platform-specific plugin."""
    if platform in SHOPIFY_PLATFORMS:
        print(f"  → Generic Shopify plugin")
        return generic_shopify.process_one(
            vendor_url, sku="", upc=upc, product_name=product_name,
            folder=OUTPUT_DIR, vendor_name=vendor_name, ptid=ptid
        )
    elif platform in MAGENTO_PLATFORMS:
        print(f"  → Generic Magento plugin")
        return generic_magento.process_one(
            vendor_url, upc=upc, product_name=product_name,
            folder=OUTPUT_DIR, vendor_name=vendor_name, ptid=ptid
        )
    elif platform in WORDPRESS_PLATFORMS or platform == "drupal":
        print(f"  → Generic WordPress plugin")
        return generic_wordpress.process_one(
            vendor_url, upc=upc, product_name=product_name,
            folder=OUTPUT_DIR, vendor_name=vendor_name, ptid=ptid,
            product_url=product_url
        )
    else:
        print(f"  → Unknown — trying Shopify then WordPress...")
        count = generic_shopify.process_one(
            vendor_url, sku="", upc=upc, product_name=product_name,
            folder=OUTPUT_DIR, vendor_name=vendor_name, ptid=ptid
        )
        if count > 0:
            return count
        return generic_wordpress.process_one(
            vendor_url, upc=upc, product_name=product_name,
            folder=OUTPUT_DIR, vendor_name=vendor_name, ptid=ptid,
            product_url=product_url
        )


# ── Excel processing ──────────────────────────────────────────────────────────
def process_excel(filepath, vendor_filter=None):
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    wb = load_workbook(filepath)
    if EXCEL_SHEET not in wb.sheetnames:
        print(f"ERROR: Sheet '{EXCEL_SHEET}' not found")
        sys.exit(1)

    ws   = wb[EXCEL_SHEET]
    rows = []

    # Row 3 = header, data starts row 4
    # Columns: A=#, B=UPC, C=PTID, D=Product Name, E=Vendor Name, F=Vendor URL, G=Status
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        upc    = str(row[1].value).strip() if row[1].value else ""  # col B
        ptid   = str(row[2].value).strip() if row[2].value else ""  # col C
        name   = str(row[3].value).strip() if row[3].value else ""  # col D
        vendor = str(row[4].value).strip() if row[4].value else ""  # col E
        url    = str(row[5].value).strip() if row[5].value else ""  # col F
        status = str(row[6].value).strip() if row[6].value else ""  # col G

        if not upc or upc == "UPC *":
            continue
        if status == "Done":
            continue
        if not url:
            continue
        if vendor_filter and vendor_filter.lower() not in vendor.lower():
            continue

        rows.append((upc, ptid, name, vendor, url, row[6]))

    if not rows:
        print("No pending products found.")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    total   = len(rows)
    success = failed = 0

    print(f"\n{'='*62}")
    print(f"  Auto Image Downloader")
    print(f"  {total} product(s) to process")
    print(f"{'='*62}\n")

    for i, (upc, ptid, name, vendor, url, status_cell) in enumerate(rows, 1):
        print(f"[{i}/{total}] {name}")
        print(f"  PTID: {ptid}  UPC: {upc}  Vendor: {vendor}")
        print(f"  URL: {url}")

        try:
            count = process_one(url, upc, name, vendor, ptid or upc)
            if count > 0:
                status_cell.value = "Done"
                success += 1
                print(f"  ✓ {count} image(s) saved\n")
            else:
                status_cell.value = "Failed"
                failed += 1
                print(f"  ✗ No images saved\n")
        except Exception as e:
            print(f"  ✗ Error: {e}\n")
            status_cell.value = "Failed"
            failed += 1

        wb.save(filepath)

    print(f"{'='*62}")
    print(f"  DONE  ✓ {success} succeeded  ✗ {failed} failed")
    print(f"  Output: {os.path.abspath(OUTPUT_DIR)}")
    print(f"{'='*62}\n")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Auto Image Downloader — detects platform, downloads images",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  Full Excel run:
      python auto_download.py --input "Product_Image_Resize_.xlsx"

  Single product:
      python auto_download.py --url https://synevit.com --upc 860011318019 --ptid 223140 --name "Neurocomplex-B Slow-Release" --vendor "Synevit"

  One vendor only:
      python auto_download.py --input "Product_Image_Resize_.xlsx" --vendor "Synevit"
        """
    )
    parser.add_argument("--input",       help="Path to Excel file")
    parser.add_argument("--url",         help="Single vendor URL")
    parser.add_argument("--upc",         default="", help="Product UPC")
    parser.add_argument("--ptid",        default="", help="Product PTID (output filename)")
    parser.add_argument("--name",        default="", help="Product name")
    parser.add_argument("--vendor",      default="Unknown", help="Vendor name — also used as output subfolder")
    parser.add_argument("--platform",    default="", help="Force platform: shopify | wordpress | woocommerce | magento")
    parser.add_argument("--product-url", default="", help="Direct product page URL — skips search/crawl")
    args = parser.parse_args()

    if args.url:
        # Single product mode
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        count = process_one(args.url, args.upc, args.name, args.vendor, args.ptid or args.upc,
                            force_platform=args.platform or None,
                            product_url=args.product_url or None)
        print(f"\nResult: {count} image(s) downloaded")
        sys.exit(0 if count > 0 else 1)

    elif args.input:
        process_excel(args.input, vendor_filter=args.vendor if args.vendor != "Unknown" else None)

    else:
        parser.print_help()
        sys.exit(1)
