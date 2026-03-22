"""
Eclectic Herb Image Downloader — Plugin for download_webp.py
=============================================================
Phase 1: Crawlee search by UPC to find product slug
Phase 2: Shopify JSON API to get images
  - Main: image filename contains 1_{size}_ matched from variant title
  - SF:   image filename starts with 3_ → crop left 22%

Usage:
    python eclecticherb_downloader.py --input "Product_Image_Resize_.xlsx"
    python eclecticherb_downloader.py --sku 3093100 --upc 023363309313 --folder resized_images/3093100

Install:
    pip install "crawlee[playwright]" requests openpyxl pillow numpy rembg onnxruntime
"""

import os, sys, asyncio, argparse, requests, shutil, tempfile, re
from io import BytesIO
from openpyxl import load_workbook
from PIL import Image, ImageFilter, ImageEnhance
import numpy as np

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resized_images")
EH_DOMAIN  = "eclecticherb.com"

SIZES = {
    "500x500_MasterDB":   (500,  500),
    "1000x1000_LiveSite": (1000, 1000),
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Accept": "application/json",
}


# ── Background removal ────────────────────────────────────────────────────────
def has_clean_background(img, threshold=0.50):
    arr   = np.array(img.convert("RGB"))
    white = (arr[:,:,0] > 235) & (arr[:,:,1] > 235) & (arr[:,:,2] > 235)
    return white.sum() / white.size > threshold

def remove_background(img):
    if has_clean_background(img):
        print(f"    [BG] Clean background — skipping rembg")
        return img.convert("RGBA")
    print(f"    [BG] Removing background...")
    try:
        from rembg import remove as rembg_remove, new_session
        session = new_session("u2net")
        buf = BytesIO()
        img.convert("RGB").save(buf, format="PNG")
        result = rembg_remove(buf.getvalue(), session=session, alpha_matting=False)
        return Image.open(BytesIO(result)).convert("RGBA")
    except Exception as e:
        print(f"    [BG] Failed ({e}) — using original")
        return img.convert("RGBA")


# ── Crop left banner (for SF images) ─────────────────────────────────────────
def crop_left_banner(img):
    """Crop the dark left banner — scan each column for dark colored pixels."""
    arr = np.array(img.convert("RGB"))
    h, w = arr.shape[:2]

    # Scan columns from left, find where banner ends
    # Banner columns have many dark/colored pixels (non-white)
    found_x = int(w * 0.22)  # safe default
    for x in range(0, int(w * 0.40)):
        col = arr[:, x, :]
        # White/light pixels: all channels > 200
        white_pixels = np.sum((col[:, 0] > 200) & (col[:, 1] > 200) & (col[:, 2] > 200))
        white_ratio  = white_pixels / h
        if white_ratio > 0.85:
            found_x = x
            break

    print(f"    [Crop] Removing left banner — cropping from x={found_x} of {w}")
    return img.crop((found_x, 0, w, h))


# ── Process and save image ────────────────────────────────────────────────────
def process_image_bytes(img_bytes, folder, name, img_type):
    try:
        img = Image.open(BytesIO(img_bytes))
        if img.width < 100 or img.height < 100:
            print(f"    Too small ({img.width}x{img.height}) — skipping")
            return False

        print(f"    Downloaded: {img.width}x{img.height}px")

        if img_type == "sfp":
            # Crop left banner then skip rembg (white background)
            img = crop_left_banner(img)
            img = img.convert("RGBA")
            print(f"    [BG] Skipping background removal for SF")
        else:
            # Main bottle — remove background
            img = remove_background(img)

        arr  = np.array(img)
        mask = arr[:,:,3] > 10
        rows = np.any(mask, axis=1)
        cols = np.any(mask, axis=0)
        if rows.any() and cols.any():
            rmin = int(np.where(rows)[0][0]);  rmax = int(np.where(rows)[0][-1])
            cmin = int(np.where(cols)[0][0]);  cmax = int(np.where(cols)[0][-1])
            ph = max(4, int((rmax - rmin) * 0.02))
            pw = max(4, int((cmax - cmin) * 0.02))
            img = img.crop((max(0,cmin-pw), max(0,rmin-ph),
                            min(img.width,cmax+pw+1), min(img.height,rmax+ph+1)))
            print(f"    [Crop] {img.width}x{img.height}")

        os.makedirs(folder, exist_ok=True)
        for label, (w, h) in SIZES.items():
            scale   = min(int(w*0.90)/img.width, int(h*0.90)/img.height)
            new_w   = int(img.width  * scale)
            new_h   = int(img.height * scale)
            resized = img.resize((new_w, new_h), Image.LANCZOS)

            rgb = resized.convert("RGB")
            rgb = rgb.filter(ImageFilter.UnsharpMask(radius=1.0, percent=120, threshold=3))
            rgb = ImageEnhance.Sharpness(rgb).enhance(1.3)
            r2,g2,b2 = rgb.split(); _,_,_,a = resized.split()
            resized = Image.merge("RGBA",(r2,g2,b2,a))

            canvas = Image.new("RGBA",(w,h),(255,255,255,255))
            canvas.paste(resized,((w-new_w)//2,(h-new_h)//2),resized)
            out = os.path.join(folder, f"{name}_{label}.jpg")
            canvas.convert("RGB").save(out, "JPEG", quality=95, dpi=(96,96))
            print(f"      → {os.path.basename(out)}")

        return True
    except Exception as e:
        print(f"    Error processing image: {e}")
        return False


# ── Step 1: Search for slug via Crawlee ───────────────────────────────────────
async def _search_one(upc):
    from crawlee.crawlers import PlaywrightCrawler, PlaywrightCrawlingContext
    from crawlee.configuration import Configuration

    search_url  = f"https://eclecticherb.com/search?q={upc}&type=product"
    slug_result = {}

    tmp_dir = tempfile.mkdtemp(prefix="eh_crawlee_")
    config  = Configuration(storage_dir=tmp_dir)

    crawler = PlaywrightCrawler(
        headless=True,
        browser_type="chromium",
        max_requests_per_crawl=2,
        configuration=config,
    )

    @crawler.router.default_handler
    async def handler(context: PlaywrightCrawlingContext) -> None:
        try:
            await context.page.wait_for_load_state("domcontentloaded")
            selectors = [
                ".product-item a[href*='/products/']",
                ".product-card a[href*='/products/']",
                "li.grid__item a[href*='/products/']",
                ".grid__item a[href*='/products/']",
                "a[href*='/products/']",
            ]
            for selector in selectors:
                links = await context.page.query_selector_all(selector)
                for link in links:
                    href = await link.get_attribute("href")
                    if not href: continue
                    if any(x in href for x in ["/search", "/collections", "?q=", "#"]): continue
                    slug = href.split("/products/")[-1].split("?")[0].split("/")[0]
                    if slug:
                        slug_result["slug"] = slug
                        break
                if "slug" in slug_result:
                    break
        except Exception as e:
            print(f"    Search error: {e}")

    await crawler.run([search_url])
    try: shutil.rmtree(tmp_dir)
    except: pass
    return slug_result.get("slug")

def search_upc(upc):
    return asyncio.run(_search_one(upc))


# ── Step 2: Get images from JSON API ─────────────────────────────────────────
def get_images_from_json(slug, upc):
    url = f"https://eclecticherb.com/products/{slug}.json"
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            print(f"    JSON API {r.status_code} for {slug}")
            return []

        data     = r.json().get("product", {})
        variants = data.get("variants", [])
        images   = data.get("images", [])

        found = []
        sorted_images = sorted(images, key=lambda x: x.get("position", 99))

        # Find variant ID matching UPC (barcode or sku field)
        matched_variant_id = None
        for variant in variants:
            barcode = str(variant.get("barcode", "")).strip()
            sku_val = str(variant.get("sku", "")).strip()
            if barcode == str(upc).strip() or sku_val == str(upc).strip():
                matched_variant_id = variant.get("id")
                break

        # Main: image whose variant_ids contains matched variant, or position 1
        main_found = False
        if matched_variant_id:
            for img in sorted_images:
                if matched_variant_id in img.get("variant_ids", []):
                    src   = img["src"]
                    fname = src.split("/")[-1].split("?")[0]
                    clean_url = src.split("?")[0] + "?width=2000"
                    found.append((clean_url, fname, "main"))
                    print(f"    Found [main]: {fname}")
                    main_found = True
                    break

        if not main_found and sorted_images:
            src   = sorted_images[0]["src"]
            fname = src.split("/")[-1].split("?")[0]
            clean_url = src.split("?")[0] + "?width=2000"
            found.append((clean_url, fname, "main"))
            print(f"    Found [main]: {fname}")

        # SF: first image with empty variant_ids (not tied to any variant)
        main_src = found[0][0] if found else None
        for img in sorted_images:
            src   = img["src"]
            fname = src.split("/")[-1].split("?")[0]
            clean_url = src.split("?")[0] + "?width=2000"
            if clean_url == main_src:
                continue  # skip main image
            if img.get("variant_ids") == []:
                found.append((clean_url, fname, "sfp"))
                print(f"    Found [sfp]: {fname}")
                break

        return found

    except Exception as e:
        print(f"    JSON error: {e}")
        return []


# ── Step 3: Download image ────────────────────────────────────────────────────
def download_image(img_url):
    try:
        r = requests.get(img_url, headers=HEADERS, timeout=20)
        if r.status_code == 200 and len(r.content) > 1000:
            return r.content
        print(f"    HTTP {r.status_code}")
        return None
    except Exception as e:
        print(f"    Download error: {e}")
        return None


# ── Main: one product at a time ───────────────────────────────────────────────
def process_one(sku, upc, folder):
    slug = search_upc(upc)
    if not slug:
        print(f"  UPC {upc} — not found on Eclectic Herb")
        return 0

    print(f"  UPC {upc} → {slug}")

    images = get_images_from_json(slug, upc)
    if not images:
        print(f"  UPC {upc} — no images found")
        return 0

    count = 0
    for img_url, fname, img_type in images:
        name      = f"{upc}_{img_type}"
        print(f"  Downloading {fname} → {name}...")
        img_bytes = download_image(img_url)
        if img_bytes and process_image_bytes(img_bytes, folder, name, img_type):
            count += 1

    return count


# ── Excel processing ──────────────────────────────────────────────────────────
def process_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb["Image Resize Jobs"]

    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        upc    = str(row[1].value).strip() if row[1].value else ""
        sku    = str(row[2].value).strip() if row[2].value else ""
        url    = str(row[5].value).strip() if row[5].value else ""
        status = str(row[6].value).strip() if row[6].value else ""

        if not upc or upc == "UPC *": continue
        if status == "Done":
            print(f"Skipping {sku} — already Done"); continue
        if not url or EH_DOMAIN not in url: continue
        if not upc:
            print(f"Skipping SKU {sku} — no UPC"); continue
        rows.append((sku, upc, row[6]))

    if not rows:
        print("No Eclectic Herb products found."); return

    total = len(rows)
    print(f"\nFound {total} Eclectic Herb products to process\n")

    for i, (sku, upc, status_cell) in enumerate(rows, 1):
        folder = os.path.join(OUTPUT_DIR, sku)
        os.makedirs(folder, exist_ok=True)
        print(f"{'='*60}")
        print(f"[{i}/{total}] SKU: {sku} | UPC: {upc}")
        print(f"{'='*60}")

        count = process_one(sku, upc, folder)
        status_cell.value = "Done" if count > 0 else "Failed"
        print(f"  → {count} images saved\n")
        wb.save(filepath)

    print(f"\nComplete!")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Eclectic Herb Image Downloader")
    parser.add_argument("--input",  help="Path to Excel file")
    parser.add_argument("--sku",    help="Single product SKU")
    parser.add_argument("--upc",    help="Single product UPC")
    parser.add_argument("--name",   help="Product name (ignored by this plugin)")
    parser.add_argument("--folder", help="Output folder for single product")
    args = parser.parse_args()

    if args.sku and args.folder:
        os.makedirs(args.folder, exist_ok=True)
        upc   = args.upc if args.upc else args.sku
        count = process_one(args.sku, upc, args.folder)
        print(f"  SKU {args.sku}: {count} images downloaded")
        sys.exit(0 if count > 0 else 1)

    elif args.input:
        if not os.path.exists(args.input):
            print(f"File not found: {args.input}"); sys.exit(1)
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        process_excel(args.input)

    else:
        print("Usage: eclecticherb_downloader.py --input file.xlsx")
        print("   or: eclecticherb_downloader.py --sku 3093100 --upc 023363309313 --folder resized_images/3093100")
        sys.exit(1)
