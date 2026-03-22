"""
Nature's Way Image Downloader — Plugin for download_webp.py
============================================================
- Search by UPC via plain requests (Shopify HTML search)
- Shopify JSON API — match images by alt tag keywords:
  PRIMARY, MAIN1, MAIN2, MAIN3, MAIN4 (back/SF)
- No background removal — clean white/transparent backgrounds

Usage:
    python naturesway_downloader.py --sku 5526 --upc 763948055265 --folder resized_images/5526
    python naturesway_downloader.py --input "Product_Image_Resize_.xlsx"
"""

import os, sys, argparse, requests, re
from io import BytesIO
from openpyxl import load_workbook
from PIL import Image, ImageFilter, ImageEnhance
import numpy as np

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resized_images")
NW_DOMAIN  = "naturesway.com"

SIZES = {
    "500x500_MasterDB":   (500,  500),
    "1000x1000_LiveSite": (1000, 1000),
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# Alt description keywords to keep
KEEP_ALT_DESC = ["primary", "right side", "left side", "back of pack"]


# ── Background removal ────────────────────────────────────────────────────────
def has_clean_background(img, threshold=0.50):
    arr   = np.array(img.convert("RGB"))
    white = (arr[:,:,0] > 235) & (arr[:,:,1] > 235) & (arr[:,:,2] > 235)
    return white.sum() / white.size > threshold

def remove_background(img):
    if has_clean_background(img):
        print(f"    [BG] Clean background — skipping rembg")
        return img.convert("RGBA")
    print(f"    [BG] Removing background...")
    try:
        from rembg import remove as rembg_remove, new_session
        session = new_session("u2net")
        buf = BytesIO()
        img.convert("RGB").save(buf, format="PNG")
        result = rembg_remove(buf.getvalue(), session=session, alpha_matting=False)
        return Image.open(BytesIO(result)).convert("RGBA")
    except Exception as e:
        print(f"    [BG] Failed ({e}) — using original")
        return img.convert("RGBA")


# ── Process and save image ────────────────────────────────────────────────────
def process_image_bytes(img_bytes, folder, name):
    try:
        img = Image.open(BytesIO(img_bytes))
        if img.width < 100 or img.height < 100:
            print(f"    Too small ({img.width}x{img.height}) — skipping")
            return False

        print(f"    Downloaded: {img.width}x{img.height}px")
        img = img.convert("RGBA")

        arr  = np.array(img)
        mask = arr[:,:,3] > 10
        rows = np.any(mask, axis=1)
        cols = np.any(mask, axis=0)
        if rows.any() and cols.any():
            rmin = int(np.where(rows)[0][0]);  rmax = int(np.where(rows)[0][-1])
            cmin = int(np.where(cols)[0][0]);  cmax = int(np.where(cols)[0][-1])
            ph = max(4, int((rmax - rmin) * 0.02))
            pw = max(4, int((cmax - cmin) * 0.02))
            img = img.crop((max(0,cmin-pw), max(0,rmin-ph),
                            min(img.width,cmax+pw+1), min(img.height,rmax+ph+1)))
            print(f"    [Crop] {img.width}x{img.height}")

        os.makedirs(folder, exist_ok=True)
        for label, (w, h) in SIZES.items():
            scale   = min(int(w*0.90)/img.width, int(h*0.90)/img.height)
            new_w   = int(img.width  * scale)
            new_h   = int(img.height * scale)
            resized = img.resize((new_w, new_h), Image.LANCZOS)

            rgb = resized.convert("RGB")
            rgb = rgb.filter(ImageFilter.UnsharpMask(radius=1.0, percent=120, threshold=3))
            rgb = ImageEnhance.Sharpness(rgb).enhance(1.3)
            r2,g2,b2 = rgb.split(); _,_,_,a = resized.split()
            resized = Image.merge("RGBA",(r2,g2,b2,a))

            canvas = Image.new("RGBA",(w,h),(255,255,255,255))
            canvas.paste(resized,((w-new_w)//2,(h-new_h)//2),resized)
            out = os.path.join(folder, f"{name}_{label}.jpg")
            canvas.convert("RGB").save(out, "JPEG", quality=95, dpi=(96,96))
            print(f"      → {os.path.basename(out)}")

        return True
    except Exception as e:
        print(f"    Error processing image: {e}")
        return False


# ── Search for slug via plain requests ───────────────────────────────────────
def search_slug(upc):
    url = f"https://naturesway.com/search?q={upc}&type=product"
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            print(f"    Search HTTP {r.status_code}")
            return None
        matches = re.findall(r'/products/([a-z0-9][a-z0-9\-]+)', r.text)
        skip    = {"all", "new", "sale", "featured"}
        for slug in matches:
            if slug not in skip and len(slug) > 3:
                print(f"    Found slug: {slug}")
                return slug
        print(f"    No product found for UPC {upc}")
        return None
    except Exception as e:
        print(f"    Search error: {e}")
        return None


# ── Get images from JSON API ──────────────────────────────────────────────────
def get_images_from_json(slug, upc):
    url = f"https://naturesway.com/products/{slug}.json"
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            print(f"    JSON API {r.status_code}")
            return []

        data   = r.json().get("product", {})
        images = data.get("images", [])
        found  = []

        for img in images:
            alt       = img.get("alt", "") or ""
            alt_upper = alt.upper()
            src       = img["src"]
            fname     = src.split("/")[-1].split("?")[0]
            clean_url = src.split("?")[0] + "?width=2000"

            # Keep only images matching desired alt descriptions
            alt_lower = alt.lower()
            matched_tag = None
            if "primary" in alt_lower:
                matched_tag = "primary"
            elif "back of pack" in alt_lower:
                matched_tag = "back"
            elif "right side" in alt_lower:
                matched_tag = "right"
            elif "left side" in alt_lower:
                matched_tag = "left"

            if matched_tag:
                found.append((clean_url, fname, matched_tag))
                print(f"    Found [{matched_tag}]: {fname}")

        # Sort: primary first, then right, left, back
        order = {"primary": 0, "right": 1, "left": 2, "back": 3}
        found.sort(key=lambda x: order.get(x[2], 99))
        return found

    except Exception as e:
        print(f"    JSON error: {e}")
        return []


# ── Download image ────────────────────────────────────────────────────────────
def download_image(img_url):
    try:
        r = requests.get(img_url, headers=HEADERS, timeout=20)
        if r.status_code == 200 and len(r.content) > 1000:
            return r.content
        print(f"    HTTP {r.status_code}")
        return None
    except Exception as e:
        print(f"    Download error: {e}")
        return None


# ── Main: one product ─────────────────────────────────────────────────────────
def process_one(sku, upc, folder):
    slug = search_slug(upc)
    if not slug:
        print(f"  UPC {upc} — not found on Nature's Way")
        return 0

    print(f"  UPC {upc} → {slug}")
    images = get_images_from_json(slug, upc)
    if not images:
        print(f"  UPC {upc} — no images found")
        return 0

    count = 0
    for img_url, fname, img_type in images:
        name      = f"{upc}_{img_type}"
        print(f"  Downloading [{img_type}] {fname}...")
        img_bytes = download_image(img_url)
        if img_bytes and process_image_bytes(img_bytes, folder, name, ):
            count += 1

    return count


# ── Excel processing ──────────────────────────────────────────────────────────
def process_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb["Image Resize Jobs"]

    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        upc    = str(row[1].value).strip() if row[1].value else ""
        sku    = str(row[2].value).strip() if row[2].value else ""
        url    = str(row[5].value).strip() if row[5].value else ""
        status = str(row[6].value).strip() if row[6].value else ""

        if not upc or upc == "UPC *": continue
        if status == "Done":
            print(f"Skipping {sku} — already Done"); continue
        if not url or NW_DOMAIN not in url: continue
        rows.append((sku, upc, row[6]))

    if not rows:
        print("No Nature's Way products found."); return

    total = len(rows)
    print(f"\nFound {total} Nature's Way products to process\n")

    for i, (sku, upc, status_cell) in enumerate(rows, 1):
        folder = os.path.join(OUTPUT_DIR, sku)
        os.makedirs(folder, exist_ok=True)
        print(f"{'='*60}")
        print(f"[{i}/{total}] SKU: {sku} | UPC: {upc}")
        print(f"{'='*60}")

        count = process_one(sku, upc, folder)
        status_cell.value = "Done" if count > 0 else "Failed"
        print(f"  → {count} images saved\n")
        wb.save(filepath)

    print(f"\nComplete!")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Nature's Way Image Downloader")
    parser.add_argument("--input",  help="Path to Excel file")
    parser.add_argument("--sku",    help="Single product SKU")
    parser.add_argument("--upc",    help="Single product UPC")
    parser.add_argument("--name",   help="Product name (ignored by this plugin)")
    parser.add_argument("--folder", help="Output folder for single product")
    args = parser.parse_args()

    if args.sku and args.folder:
        os.makedirs(args.folder, exist_ok=True)
        upc   = args.upc if args.upc else args.sku
        count = process_one(args.sku, upc, args.folder)
        print(f"  SKU {args.sku}: {count} images downloaded")
        sys.exit(0 if count > 0 else 1)

    elif args.input:
        if not os.path.exists(args.input):
            print(f"File not found: {args.input}"); sys.exit(1)
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        process_excel(args.input)

    else:
        print("Usage: naturesway_downloader.py --input file.xlsx")
        print("   or: naturesway_downloader.py --sku 5526 --upc 763948055265 --folder resized_images/5526")
        sys.exit(1)
