"""
Product Image Download Pipeline — Orchestrator
================================================
Reads Product_Image_Resize_.xlsx row by row, detects vendor by URL domain,
runs the correct plugin script, marks row Done/Failed in the Excel.

HOW TO ADD A NEW VENDOR
-----------------------
1. Create <vendor>_downloader.py in the same folder.
   Must accept:  --sku  --upc  --name  --folder
   Must exit 0 on success, non-zero on failure.

2. Add ONE line to VENDOR_REGISTRY below:
       "domain.com": ("script.py", "Display Name", timeout_seconds),

Nothing else needs to change.

Usage:
    python download_webp.py --input "Product_Image_Resize_.xlsx"
"""

import os
import sys
import argparse
import subprocess
from urllib.parse import urlparse
from openpyxl import load_workbook


# ══════════════════════════════════════════════════════════════════════════════
# VENDOR REGISTRY
# Add a new vendor by adding ONE line here + dropping the plugin script.
# Format: "domain.com": ("script.py", "Display Name", timeout_seconds),
# ══════════════════════════════════════════════════════════════════════════════
VENDOR_REGISTRY = {
    "herb-pharm.com":                ("herbpharm_downloader.py",     "Herb Pharm",          180),
    "naturesway.com":                ("naturesway_downloader.py",    "Nature's Way",        180),
    "irwinnaturals.com":             ("irwinnaturals_downloader.py", "Irwin Naturals",      180),
    "gardenoflife.com":              ("gol_downloader.py",           "Garden of Life",      180),
    "newchapter.com":                ("newchapter_downloader.py",    "New Chapter",         180),
    "hostdefense.com":               ("hostdefense_downloader.py",   "Host Defense",        180),
    "traceminerals.com":             ("traceminerals_downloader.py", "Trace Minerals",      180),
    "eclecticherb.com":              ("eclecticherb_downloader.py",  "Eclectic Herb",       180),
    "jarrow.com":                    ("jarrow_downloader.py",        "Jarrow Formulas",     180),
    "lifeextension.com":             ("lifeextension_downloader.py", "Life Extension",      120),
    "puritan.com":                   ("puritan_downloader.py",       "Puritan's Pride",     120),
    "vitapluslifelineusa.com":       ("vitaplus_downloader.py",      "Vita Plus",           120),
    "naturally.com":                 ("naturally_downloader.py",     "Naturally Vitamins",  120),
    "megafood.com":                  ("megafood_downloader.py",      "MegaFood",            180),
    "nowfoods.com":                  ("nowfoods_downloader.py",      "Now Foods",           180),
    "northamericanherbandspice.com": ("nahs_downloader.py",          "NAHS",                120),
    "drbvitamins.com":               ("doctorsbest_downloader.py",   "Doctor's Best",       180),
    "doctorsbest.com":               ("doctorsbest_downloader.py",   "Doctor's Best",       180),
    "nutricost.com":                 ("nutricost_downloader.py",     "Nutricost",           180),
    "countrylifevitamins.com":       ("countrylife_downloader.py",   "Country Life",        180),
    # Uncomment when plugin is ready:
    # "nutricology.com":             ("nutricology_downloader.py",   "Nutricology",         180),
    # "allergyresearchgroup.com":    ("arg_downloader.py",           "Allergy Research",    180),
}

EXCEL_SHEET = "Image Resize Jobs"
DATA_START  = 4
COL_UPC     = 1   # B
COL_SKU     = 2   # C
COL_NAME    = 3   # D
COL_VENDOR  = 4   # E
COL_URL     = 5   # F
COL_STATUS  = 6   # G

OUTPUT_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resized_images")
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))


def get_domain(url):
    try:
        return urlparse(url).netloc.lower().replace("www.", "")
    except Exception:
        return ""


def match_vendor(url):
    domain = get_domain(url)
    for registered, info in VENDOR_REGISTRY.items():
        if registered in domain:
            return info
    return None


def run_plugin(script_file, sku, upc, name, folder, timeout):
    script_path = os.path.join(SCRIPT_DIR, script_file)
    if not os.path.exists(script_path):
        print(f"  [ERROR] Plugin not found: {script_file}")
        print(f"          Place it in the same folder as download_webp.py")
        return 0
    cmd = [
        sys.executable, script_path,
        "--sku",    str(sku),
        "--upc",    str(upc),
        "--name",   str(name),
        "--folder", folder,
    ]
    try:
        result = subprocess.run(cmd, capture_output=False, timeout=timeout)
        if result.returncode == 0:
            if os.path.isdir(folder):
                jpgs = [f for f in os.listdir(folder) if f.lower().endswith(".jpg")]
                return max(len(jpgs) // 2, 1) if jpgs else 0
            return 0
        else:
            print(f"  [ERROR] Plugin exited with code {result.returncode}")
            return 0
    except subprocess.TimeoutExpired:
        print(f"  [ERROR] Timed out after {timeout}s")
        return 0
    except Exception as e:
        print(f"  [ERROR] {e}")
        return 0


def process_excel(filepath):
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    wb = load_workbook(filepath)
    if EXCEL_SHEET not in wb.sheetnames:
        print(f"ERROR: Sheet '{EXCEL_SHEET}' not found. Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[EXCEL_SHEET]
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    total = success = failed = 0

    for row in ws.iter_rows(min_row=DATA_START, max_row=ws.max_row):
        upc    = str(row[COL_UPC].value).strip()    if row[COL_UPC].value    else ""
        sku    = str(row[COL_SKU].value).strip()    if row[COL_SKU].value    else ""
        name   = str(row[COL_NAME].value).strip()   if row[COL_NAME].value   else ""
        url    = str(row[COL_URL].value).strip()    if row[COL_URL].value    else ""
        status = str(row[COL_STATUS].value).strip() if row[COL_STATUS].value else ""

        if not upc or upc == "UPC *":
            continue
        if status == "Done":
            print(f"Skipping {sku or upc} — already Done")
            continue
        if not url:
            print(f"Skipping {sku or upc} — no vendor URL")
            row[COL_STATUS].value = "Failed - No URL"
            wb.save(filepath)
            failed += 1
            continue

        total += 1
        identifier = sku if sku else upc
        folder = os.path.join(OUTPUT_DIR, identifier)

        print(f"\n{'='*60}")
        print(f"[{total}] SKU: {sku or '—'} | UPC: {upc} | {name}")
        print(f"{'='*60}")

        plugin_info = match_vendor(url)
        if not plugin_info:
            print(f"  No plugin for domain: {get_domain(url)}")
            print(f"  Add it to VENDOR_REGISTRY in download_webp.py")
            row[COL_STATUS].value = "Failed - No Plugin"
            wb.save(filepath)
            failed += 1
            continue

        script, display_name, timeout = plugin_info
        print(f"  Plugin : {display_name} ({script})")

        os.makedirs(folder, exist_ok=True)
        count = run_plugin(script, sku, upc, name, folder, timeout)

        if count > 0:
            row[COL_STATUS].value = "Done"
            success += 1
            print(f"  → {count} image set(s) saved")
        else:
            row[COL_STATUS].value = "Failed"
            failed += 1
            print(f"  → Failed")

        wb.save(filepath)

    print(f"\n{'='*60}")
    print(f"DONE! {success} succeeded, {failed} failed out of {total} total.")
    print(f"Output: {os.path.abspath(OUTPUT_DIR)}")
    print(f"{'='*60}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Product Image Download Pipeline")
    parser.add_argument("--input", required=True, help="Path to Excel file")
    args = parser.parse_args()
    process_excel(args.input)
