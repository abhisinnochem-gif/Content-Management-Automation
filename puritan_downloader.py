"""
Puritan's Pride Image Downloader — Plugin for download_webp.py
===============================================================
Uses Crawlee + Playwright to bypass 403 bot protection.
Direct CDN download using SKU — no searching needed.

URL pattern:
    puritan.com/VF/puritanspride/product_images/product_detail/{SKU6}.jpg
    puritan.com/VF/puritanspride/product_images/product_detail/{SKU6}L.jpg
    puritan.com/VF/puritanspride/product_images/product_detail/{SKU6}R.jpg

Usage:
    python puritan_downloader.py --input "Product_Image_Resize_.xlsx"
    python puritan_downloader.py --sku 10157 --upc 025077101573 --folder resized_images/10157

Install:
    pip install "crawlee[playwright]" openpyxl pillow numpy rembg onnxruntime
"""

import os, sys, asyncio, argparse, shutil, tempfile
from io import BytesIO
from openpyxl import load_workbook
from PIL import Image, ImageFilter, ImageEnhance
import numpy as np

OUTPUT_DIR     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resized_images")
PURITAN_DOMAIN = "puritan.com"
BASE_URL       = "https://www.puritan.com/VF/puritanspride/product_images/product_detail"

SIZES = {
    "500x500_MasterDB":   (500,  500),
    "1000x1000_LiveSite": (1000, 1000),
}


def pad_sku(sku):
    """Pad SKU to 6 digits: 10157 → 010157"""
    return str(sku).strip().lstrip("0").zfill(6)


# ── Background removal ────────────────────────────────────────────────────────
def has_clean_background(img, threshold=0.15):
    arr   = np.array(img.convert("RGB"))
    white = (arr[:,:,0] > 235) & (arr[:,:,1] > 235) & (arr[:,:,2] > 235)
    return white.sum() / white.size > threshold

def remove_background(img):
    if has_clean_background(img):
        print(f"    [BG] Clean background — skipping rembg")
        return img.convert("RGBA")
    print(f"    [BG] Removing background...")
    try:
        from rembg import remove as rembg_remove, new_session
        session = new_session("u2net")
        buf = BytesIO()
        img.save(buf, format="PNG")
        result = rembg_remove(buf.getvalue(), session=session, alpha_matting=False)
        return Image.open(BytesIO(result)).convert("RGBA")
    except Exception as e:
        print(f"    [BG] Failed ({e}) — using original")
        return img.convert("RGBA")


# ── Process and save image ────────────────────────────────────────────────────
def process_image_bytes(img_bytes, folder, name):
    try:
        img = Image.open(BytesIO(img_bytes))
        if img.width < 100 or img.height < 100:
            print(f"    Too small ({img.width}x{img.height}) — skipping")
            return False

        print(f"    Downloaded: {img.width}x{img.height}px")
        # Skip background removal — Puritan images already have clean backgrounds
        img = img.convert("RGBA")

        arr  = np.array(img)
        mask = arr[:,:,3] > 10
        rows = np.any(mask, axis=1)
        cols = np.any(mask, axis=0)
        if rows.any() and cols.any():
            rmin = int(np.where(rows)[0][0]);  rmax = int(np.where(rows)[0][-1])
            cmin = int(np.where(cols)[0][0]);  cmax = int(np.where(cols)[0][-1])
            ph = max(4, int((rmax - rmin) * 0.02))
            pw = max(4, int((cmax - cmin) * 0.02))
            img = img.crop((max(0,cmin-pw), max(0,rmin-ph),
                            min(img.width,cmax+pw+1), min(img.height,rmax+ph+1)))
            print(f"    [Crop] {img.width}x{img.height}")

        os.makedirs(folder, exist_ok=True)
        for label, (w, h) in SIZES.items():
            scale   = min(int(w*0.90)/img.width, int(h*0.90)/img.height)
            new_w   = int(img.width  * scale)
            new_h   = int(img.height * scale)
            resized = img.resize((new_w, new_h), Image.LANCZOS)

            rgb = resized.convert("RGB")
            rgb = rgb.filter(ImageFilter.UnsharpMask(radius=1.0, percent=120, threshold=3))
            rgb = ImageEnhance.Sharpness(rgb).enhance(1.3)
            r2,g2,b2 = rgb.split(); _,_,_,a = resized.split()
            resized = Image.merge("RGBA",(r2,g2,b2,a))

            canvas = Image.new("RGBA",(w,h),(255,255,255,255))
            canvas.paste(resized,((w-new_w)//2,(h-new_h)//2),resized)
            out = os.path.join(folder, f"{name}_{label}.jpg")
            canvas.convert("RGB").save(out, "JPEG", quality=95, dpi=(96,96))
            print(f"      → {os.path.basename(out)}")

        return True
    except Exception as e:
        print(f"    Error processing image: {e}")
        return False


# ── Download via Crawlee (bypasses 403) ──────────────────────────────────────
async def _download_images(urls):
    """Download list of URLs using Crawlee. Returns {url: bytes}"""
    from crawlee.crawlers import PlaywrightCrawler, PlaywrightCrawlingContext
    from crawlee.configuration import Configuration

    tmp_dir = tempfile.mkdtemp(prefix="puritan_crawlee_")
    config  = Configuration(storage_dir=tmp_dir)
    results = {}

    crawler = PlaywrightCrawler(
        headless=True,
        browser_type="chromium",
        max_requests_per_crawl=len(urls) + 1,
        configuration=config,
    )

    @crawler.router.default_handler
    async def handler(context: PlaywrightCrawlingContext) -> None:
        url = context.request.url
        try:
            response = await context.page.goto(url)
            if response and response.status == 200:
                body = await response.body()
                results[url] = body
                print(f"  ✓ {len(body)} bytes: {url.split('/')[-1]}")
            elif response and response.status == 404:
                print(f"  Not found (404): {url.split('/')[-1]}")
                results[url] = None
            else:
                status = response.status if response else "no response"
                print(f"  HTTP {status}: {url.split('/')[-1]}")
                results[url] = None
        except Exception as e:
            print(f"  Error: {e}")
            results[url] = None

    await crawler.run(urls)
    try: shutil.rmtree(tmp_dir)
    except: pass
    return results


def process_one(sku, upc, folder):
    sku6   = pad_sku(sku)
    images = [
        (f"{BASE_URL}/{sku6}.jpg",  f"{upc}_front"),
        (f"{BASE_URL}/{sku6}L.jpg", f"{upc}_left"),
        (f"{BASE_URL}/{sku6}R.jpg", f"{upc}_right"),
    ]

    urls     = [img[0] for img in images]
    name_map = {img[0]: img[1] for img in images}

    print(f"  Downloading 3 images via Crawlee...")
    results = asyncio.run(_download_images(urls))

    os.makedirs(folder, exist_ok=True)
    count = 0
    for url, name in images:
        data = results.get(url)
        if data and len(data) > 1000:
            if process_image_bytes(data, folder, name):
                count += 1

    return count


# ── Excel processing ──────────────────────────────────────────────────────────
def process_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb["Image Resize Jobs"]

    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        upc    = str(row[1].value).strip() if row[1].value else ""
        sku    = str(row[2].value).strip() if row[2].value else ""
        url    = str(row[5].value).strip() if row[5].value else ""
        status = str(row[6].value).strip() if row[6].value else ""

        if not upc or upc == "UPC *": continue
        if status == "Done":
            print(f"Skipping {sku} — already Done"); continue
        if not url or PURITAN_DOMAIN not in url: continue
        if not sku:
            print(f"Skipping UPC {upc} — no SKU"); continue
        rows.append((sku, upc, row[6]))

    if not rows:
        print("No Puritan's Pride products found."); return

    total = len(rows)
    print(f"\nFound {total} Puritan's Pride products to process\n")

    for i, (sku, upc, status_cell) in enumerate(rows, 1):
        folder = os.path.join(OUTPUT_DIR, sku)
        os.makedirs(folder, exist_ok=True)
        print(f"{'='*60}")
        print(f"[{i}/{total}] SKU: {sku} | UPC: {upc}")
        print(f"{'='*60}")

        count = process_one(sku, upc, folder)
        status_cell.value = "Done" if count > 0 else "Failed"
        print(f"  → {count} images saved\n")
        wb.save(filepath)

    print(f"\nComplete!")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Puritan's Pride Image Downloader")
    parser.add_argument("--input",  help="Path to Excel file")
    parser.add_argument("--sku",    help="Single product SKU")
    parser.add_argument("--upc",    help="Single product UPC")
    parser.add_argument("--name",   help="Product name (ignored by this plugin)")
    parser.add_argument("--folder", help="Output folder for single product")
    args = parser.parse_args()

    if args.sku and args.folder:
        os.makedirs(args.folder, exist_ok=True)
        upc   = args.upc if args.upc else args.sku
        count = process_one(args.sku, upc, args.folder)
        print(f"  SKU {args.sku}: {count} images downloaded")
        sys.exit(0 if count > 0 else 1)

    elif args.input:
        if not os.path.exists(args.input):
            print(f"File not found: {args.input}"); sys.exit(1)
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        process_excel(args.input)

    else:
        print("Usage: puritan_downloader.py --input file.xlsx")
        print("   or: puritan_downloader.py --sku 10157 --upc 025077101573 --folder resized_images/10157")
        sys.exit(1)
