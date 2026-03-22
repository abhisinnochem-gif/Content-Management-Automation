"""
Herb Pharm Image Downloader — Plugin for download_webp.py
=========================================================
- Search by SKU via Crawlee → get product slug
- Shopify JSON API:
  - Main: variant matched by UPC barcode → image_id
  - SF/Label: image filename contains 'WebLabel' + SKU → crop left portion

Usage:
    python herbpharm_downloader.py --sku DDCLAW01 --upc 090700000479 --folder resized_images/DDCLAW01
    python herbpharm_downloader.py --input "Product_Image_Resize_.xlsx"
"""

import os, sys, argparse, requests
from io import BytesIO
from openpyxl import load_workbook
from PIL import Image, ImageFilter, ImageEnhance
import numpy as np

OUTPUT_DIR     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resized_images")
HP_DOMAIN      = "herb-pharm.com"

SIZES = {
    "500x500_MasterDB":   (500,  500),
    "1000x1000_LiveSite": (1000, 1000),
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Accept": "application/json",
}


# ── Background removal ────────────────────────────────────────────────────────
def has_clean_background(img, threshold=0.50):
    arr   = np.array(img.convert("RGB"))
    white = (arr[:,:,0] > 235) & (arr[:,:,1] > 235) & (arr[:,:,2] > 235)
    return white.sum() / white.size > threshold

def remove_background(img):
    if has_clean_background(img):
        print(f"    [BG] Clean background — skipping rembg")
        return img.convert("RGBA")
    print(f"    [BG] Removing background...")
    try:
        from rembg import remove as rembg_remove, new_session
        session = new_session("u2net")
        buf = BytesIO()
        img.convert("RGB").save(buf, format="PNG")
        result = rembg_remove(buf.getvalue(), session=session, alpha_matting=False)
        return Image.open(BytesIO(result)).convert("RGBA")
    except Exception as e:
        print(f"    [BG] Failed ({e}) — using original")
        return img.convert("RGBA")


# ── Crop SF label — keep right white panel only ───────────────────────────────
def crop_sf_label(img):
    """Crop the orange left portion — keep only the white SF panel on the right."""
    arr = np.array(img.convert("RGB"))
    h, w = arr.shape[:2]

    # Fixed crop — SF panel always starts at ~55% from left
    # Scan from 55% to find first white column
    found_x = int(w * 0.55)
    for x in range(int(w * 0.55), int(w * 0.80)):
        col          = arr[:, x, :]
        white_pixels = np.sum((col[:, 0] > 210) & (col[:, 1] > 210) & (col[:, 2] > 210))
        if white_pixels / h > 0.60:
            found_x = x
            break

    # Crop bottom — find the "HERB PHARM" line by detecting a mostly-empty
    # (non-white, non-text) orange row gap before the bold address block.
    # Strategy: scan from bottom up, find last row with ANY white pixels
    # (disclaimer box is white-bordered), then find the gap above address block.
    # Simpler: find last row where >30% pixels are white (end of SF content area)
    found_y = int(h * 0.78)  # safe default ~78%
    # Scan upward from 95% to find last white-heavy row
    for y in range(int(h * 0.95), int(h * 0.50), -1):
        row          = arr[y, :, :]
        white_pixels = np.sum((row[:, 0] > 210) & (row[:, 1] > 210) & (row[:, 2] > 210))
        if white_pixels / w > 0.20:
            found_y = y + 10
            break

    # Crop right side — find last sustained white column from right
    found_x2  = w
    white_run2 = 0
    for x in range(w - 1, int(w * 0.50), -1):
        col          = arr[:, x, :]
        white_pixels = np.sum((col[:, 0] > 200) & (col[:, 1] > 200) & (col[:, 2] > 200))
        if white_pixels / h > 0.50:
            white_run2 += 1
            if white_run2 >= 10:
                found_x2 = x + 10
                break
        else:
            white_run2 = 0

    # Crop top — skip any non-white rows at top
    found_y0 = 0
    for y in range(0, int(h * 0.30)):
        row          = arr[y, :, :]
        white_pixels = np.sum((row[:, 0] > 200) & (row[:, 1] > 200) & (row[:, 2] > 200))
        if white_pixels / w > 0.30:
            found_y0 = max(0, y - 2)
            break

    print(f"    [Crop SF] x={found_x}:{found_x2}, y={found_y0}:{found_y}")
    return img.crop((found_x, found_y0, found_x2, found_y))


# ── Process and save image ────────────────────────────────────────────────────
def process_image_bytes(img_bytes, folder, name, img_type):
    try:
        img = Image.open(BytesIO(img_bytes))
        if img.width < 100 or img.height < 100:
            print(f"    Too small ({img.width}x{img.height}) — skipping")
            return False

        print(f"    Downloaded: {img.width}x{img.height}px")

        if img_type == "sfp":
            img = crop_sf_label(img)
            img = img.convert("RGBA")
            print(f"    [BG] Skipping rembg for SF label")
        else:
            img = remove_background(img)

        arr  = np.array(img)
        mask = arr[:,:,3] > 10
        rows = np.any(mask, axis=1)
        cols = np.any(mask, axis=0)
        if rows.any() and cols.any():
            rmin = int(np.where(rows)[0][0]);  rmax = int(np.where(rows)[0][-1])
            cmin = int(np.where(cols)[0][0]);  cmax = int(np.where(cols)[0][-1])
            ph = max(4, int((rmax - rmin) * 0.02))
            pw = max(4, int((cmax - cmin) * 0.02))
            img = img.crop((max(0,cmin-pw), max(0,rmin-ph),
                            min(img.width,cmax+pw+1), min(img.height,rmax+ph+1)))
            print(f"    [Crop] {img.width}x{img.height}")

        os.makedirs(folder, exist_ok=True)
        for label, (w, h) in SIZES.items():
            scale   = min(int(w*0.90)/img.width, int(h*0.90)/img.height)
            new_w   = int(img.width  * scale)
            new_h   = int(img.height * scale)
            resized = img.resize((new_w, new_h), Image.LANCZOS)

            rgb = resized.convert("RGB")
            rgb = rgb.filter(ImageFilter.UnsharpMask(radius=1.0, percent=120, threshold=3))
            rgb = ImageEnhance.Sharpness(rgb).enhance(1.3)
            r2,g2,b2 = rgb.split(); _,_,_,a = resized.split()
            resized = Image.merge("RGBA",(r2,g2,b2,a))

            canvas = Image.new("RGBA",(w,h),(255,255,255,255))
            canvas.paste(resized,((w-new_w)//2,(h-new_h)//2),resized)
            out = os.path.join(folder, f"{name}_{label}.jpg")
            canvas.convert("RGB").save(out, "JPEG", quality=95, dpi=(96,96))
            print(f"      → {os.path.basename(out)}")

        return True
    except Exception as e:
        print(f"    Error processing image: {e}")
        return False


# ── Find slug via Shopify collections API (no browser needed) ────────────────
def search_slug(sku):
    """Page through collections/all/products.json to find product with matching SKU."""
    sku_upper = sku.strip().upper()
    page = 1
    while True:
        url = f"https://www.herb-pharm.com/collections/all/products.json?limit=250&page={page}"
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            if r.status_code != 200:
                print(f"    Collections API {r.status_code}")
                break
            products = r.json().get("products", [])
            if not products:
                break
            for product in products:
                for variant in product.get("variants", []):
                    v_sku = str(variant.get("sku", "")).strip().upper()
                    if v_sku == sku_upper:
                        slug = product.get("handle", "")
                        print(f"    Found slug: {slug}")
                        return slug
            page += 1
        except Exception as e:
            print(f"    Collections API error: {e}")
            break
    print(f"    SKU {sku} not found in collections")
    return None


# ── Get images from JSON API ──────────────────────────────────────────────────
def get_images_from_json(slug, sku, upc):
    url = f"https://www.herb-pharm.com/products/{slug}.json"
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            print(f"    JSON API {r.status_code} for {slug}")
            return []

        data        = r.json().get("product", {})
        variants    = data.get("variants", [])
        images      = data.get("images", [])
        image_by_id = {img["id"]: img for img in images}
        sku_upper   = sku.strip().upper()
        found       = []

        # Main: match variant by UPC barcode → image_id
        for variant in variants:
            if str(variant.get("barcode", "")).strip() == str(upc).strip():
                image_id = variant.get("image_id")
                if image_id and image_id in image_by_id:
                    src       = image_by_id[image_id]["src"]
                    fname     = src.split("/")[-1].split("?")[0]
                    clean_url = src.split("?")[0] + "?width=2000"
                    found.append((clean_url, fname, "main"))
                    print(f"    Found [main]: {fname}")
                break

        # SF: image filename contains 'WebLabel' and SKU
        for img in images:
            src   = img["src"]
            fname = src.split("/")[-1].split("?")[0]
            fname_upper = fname.upper()
            if "WEBLABEL" in fname_upper and sku_upper in fname_upper:
                clean_url = src.split("?")[0] + "?width=2000"
                found.append((clean_url, fname, "sfp"))
                print(f"    Found [sfp]: {fname}")
                break

        # SF fallback: any WebLabel image if SKU not in filename
        if len(found) == 1 and found[0][2] == "main":
            for img in images:
                src   = img["src"]
                fname = src.split("/")[-1].split("?")[0]
                if "WEBLABEL" in fname.upper():
                    clean_url = src.split("?")[0] + "?width=2000"
                    found.append((clean_url, fname, "sfp"))
                    print(f"    Found [sfp fallback]: {fname}")
                    break

        return found

    except Exception as e:
        print(f"    JSON error: {e}")
        return []


# ── Download image ────────────────────────────────────────────────────────────
def download_image(img_url):
    try:
        r = requests.get(img_url, headers=HEADERS, timeout=20)
        if r.status_code == 200 and len(r.content) > 1000:
            return r.content
        print(f"    HTTP {r.status_code}")
        return None
    except Exception as e:
        print(f"    Download error: {e}")
        return None


# ── Main: one product ─────────────────────────────────────────────────────────
def process_one(sku, upc, folder):
    slug = search_slug(sku)
    if not slug:
        print(f"  SKU {sku} — not found on Herb Pharm")
        return 0

    print(f"  SKU {sku} → {slug}")
    images = get_images_from_json(slug, sku, upc)
    if not images:
        print(f"  SKU {sku} — no images found")
        return 0

    count = 0
    for img_url, fname, img_type in images:
        name      = f"{upc}_{img_type}"
        print(f"  Downloading [{img_type}] {fname}...")
        img_bytes = download_image(img_url)
        if img_bytes and process_image_bytes(img_bytes, folder, name, img_type):
            count += 1

    return count


# ── Excel processing ──────────────────────────────────────────────────────────
def process_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb["Image Resize Jobs"]

    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        upc    = str(row[1].value).strip() if row[1].value else ""
        sku    = str(row[2].value).strip() if row[2].value else ""
        url    = str(row[5].value).strip() if row[5].value else ""
        status = str(row[6].value).strip() if row[6].value else ""

        if not upc or upc == "UPC *": continue
        if status == "Done":
            print(f"Skipping {sku} — already Done"); continue
        if not url or HP_DOMAIN not in url: continue
        rows.append((sku, upc, row[6]))

    if not rows:
        print("No Herb Pharm products found."); return

    total = len(rows)
    print(f"\nFound {total} Herb Pharm products to process\n")

    for i, (sku, upc, status_cell) in enumerate(rows, 1):
        folder = os.path.join(OUTPUT_DIR, sku)
        os.makedirs(folder, exist_ok=True)
        print(f"{'='*60}")
        print(f"[{i}/{total}] SKU: {sku} | UPC: {upc}")
        print(f"{'='*60}")

        count = process_one(sku, upc, folder)
        status_cell.value = "Done" if count > 0 else "Failed"
        print(f"  → {count} images saved\n")
        wb.save(filepath)

    print(f"\nComplete!")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Herb Pharm Image Downloader")
    parser.add_argument("--input",  help="Path to Excel file")
    parser.add_argument("--sku",    help="Single product SKU")
    parser.add_argument("--upc",    help="Single product UPC")
    parser.add_argument("--name",   help="Product name (ignored by this plugin)")
    parser.add_argument("--folder", help="Output folder for single product")
    args = parser.parse_args()

    if args.sku and args.folder:
        os.makedirs(args.folder, exist_ok=True)
        upc   = args.upc if args.upc else args.sku
        count = process_one(args.sku, upc, args.folder)
        print(f"  SKU {args.sku}: {count} images downloaded")
        sys.exit(0 if count > 0 else 1)

    elif args.input:
        if not os.path.exists(args.input):
            print(f"File not found: {args.input}"); sys.exit(1)
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        process_excel(args.input)

    else:
        print("Usage: herbpharm_downloader.py --input file.xlsx")
        print("   or: herbpharm_downloader.py --sku DDCLAW01 --upc 090700000479 --folder resized_images/DDCLAW01")
        sys.exit(1)
