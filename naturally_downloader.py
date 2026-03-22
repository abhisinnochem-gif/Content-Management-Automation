"""
Naturally Vitamins (naturally.com) Image Downloader
=====================================================
Magento store — static HTML product pages.

Key behaviour:
  - Searches by product name → finds product page URL
  - Scrapes ALL images from the product page
  - Picks the EXACT image by matching count/size from UPC or product name
    e.g. UPC 032115602503 (250ct) → picks formula50-250.jpg NOT formula50-100.jpg
  - Falls back to first image only if no count can be determined

Usage:
    python naturally_downloader.py --sku M60250 --upc 032115602503 --name "Formula 50 250 Softgels" --folder resized_images/M60250
    python naturally_downloader.py --input "Product_Image_Resize_.xlsx"

Install:
    pip install requests openpyxl beautifulsoup4 pillow numpy rembg onnxruntime
"""

import os, sys, argparse, requests, re
from io import BytesIO
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from PIL import Image, ImageFilter, ImageEnhance
import numpy as np

OUTPUT_DIR        = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resized_images")
NATURALLY_DOMAIN  = "naturally.com"
BASE_URL          = "https://naturally.com"
SEARCH_URL        = "https://naturally.com/store/catalogsearch/result/?q={query}"

SIZES = {
    "500x500_MasterDB":   (500,  500),
    "1000x1000_LiveSite": (1000, 1000),
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
}

# Typical product count values to recognise in UPC / product name
KNOWN_COUNTS = {30,45,60,90,100,120,150,180,200,240,250,300,360,365,500,1000}


# ── Extract count from product name ─────────────────────────────────────────
def extract_count(product_name):
    """
    The count is ALWAYS the last/unit number in the product name.
    Checks for number before unit word first, then last number in name.
      "Formula 50 250 Softgels"               -> 250
      "Medizym Systemic Enzyme Formula 800 TABS" -> 800
      "Nattokinase 1500 Systemic Enzyme 120 TABS" -> 120  (not 1500)
      "Histame 30 caps"                        -> 30
    """
    # Number immediately before a unit word = the count
    m = re.search(r'\b(\d+)\s*(?:softgels?|tablets?|tabs?|capsules?|caps?|ct|count)\b', product_name, re.I)
    if m:
        return m.group(1)
    # Fallback: last standalone number in the name
    nums = re.findall(r'\b(\d+)\b', product_name)
    return nums[-1] if nums else None


# ── Background removal ────────────────────────────────────────────────────────
def has_clean_background(img, threshold=0.50):
    arr   = np.array(img.convert("RGB"))
    white = (arr[:,:,0] > 235) & (arr[:,:,1] > 235) & (arr[:,:,2] > 235)
    return white.sum() / white.size > threshold

def remove_background(img):
    if has_clean_background(img):
        print(f"    [BG] Clean background — skipping rembg")
        return img.convert("RGBA")
    print(f"    [BG] Removing background...")
    try:
        from rembg import remove as rembg_remove, new_session
        session = new_session("u2net")
        buf = BytesIO()
        img.convert("RGB").save(buf, format="PNG")
        result = rembg_remove(buf.getvalue(), session=session, alpha_matting=False)
        return Image.open(BytesIO(result)).convert("RGBA")
    except Exception as e:
        print(f"    [BG] Failed ({e}) — using original")
        return img.convert("RGBA")


# ── Process and save image ────────────────────────────────────────────────────
def process_image_bytes(img_bytes, folder, name):
    try:
        img = Image.open(BytesIO(img_bytes))
        if img.width < 100 or img.height < 100:
            print(f"    Too small ({img.width}x{img.height}) — skipping")
            return False

        print(f"    Downloaded: {img.width}x{img.height}px")
        img = remove_background(img)

        arr  = np.array(img)
        mask = arr[:,:,3] > 10
        rows = np.any(mask, axis=1)
        cols = np.any(mask, axis=0)
        if rows.any() and cols.any():
            rmin = int(np.where(rows)[0][0]);  rmax = int(np.where(rows)[0][-1])
            cmin = int(np.where(cols)[0][0]);  cmax = int(np.where(cols)[0][-1])
            ph = max(4, int((rmax - rmin) * 0.02))
            pw = max(4, int((cmax - cmin) * 0.02))
            img = img.crop((max(0,cmin-pw), max(0,rmin-ph),
                            min(img.width,cmax+pw+1), min(img.height,rmax+ph+1)))
            print(f"    [Crop] {img.width}x{img.height}")

        os.makedirs(folder, exist_ok=True)
        for label, (w, h) in SIZES.items():
            scale   = min(int(w*0.90)/img.width, int(h*0.90)/img.height)
            new_w   = int(img.width  * scale)
            new_h   = int(img.height * scale)
            resized = img.resize((new_w, new_h), Image.LANCZOS)

            rgb = resized.convert("RGB")
            rgb = rgb.filter(ImageFilter.UnsharpMask(radius=1.0, percent=120, threshold=3))
            rgb = ImageEnhance.Sharpness(rgb).enhance(1.3)
            r2,g2,b2 = rgb.split(); _,_,_,a = resized.split()
            resized = Image.merge("RGBA",(r2,g2,b2,a))

            canvas = Image.new("RGBA",(w,h),(255,255,255,255))
            canvas.paste(resized,((w-new_w)//2,(h-new_h)//2),resized)
            out = os.path.join(folder, f"{name}_{label}.jpg")
            canvas.convert("RGB").save(out, "JPEG", quality=95, dpi=(96,96))
            print(f"      → {os.path.basename(out)}")

        return True
    except Exception as e:
        print(f"    Error processing image: {e}")
        return False


# ── Step 1: Search for product page URL ──────────────────────────────────────
def search_product(product_name):
    query_words = [w for w in product_name.split() if len(w) > 2][:5]
    query       = requests.utils.quote(" ".join(query_words))
    search_url  = SEARCH_URL.format(query=query)
    print(f"    Searching: {search_url}")

    try:
        r = requests.get(search_url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            print(f"    Search HTTP {r.status_code}")
            return None

        soup       = BeautifulSoup(r.text, "html.parser")
        name_low   = product_name.lower().strip()
        best_url   = None
        best_score = 0

        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if "/store/" not in href or not href.endswith(".html"):
                continue
            if any(x in href for x in ["cart", "account", "login", "catalogsearch"]):
                continue

            text  = a.get_text(strip=True).lower()
            slug  = href.lower().split("/store/")[-1].replace(".html","").replace("-"," ").replace("/"," ")
            score = 0
            words = [w for w in name_low.split() if len(w) > 2]
            score += sum(2 for w in words if w in slug)
            score += sum(1 for w in words if w in text)
            if name_low in text:
                score += 5

            if score > best_score:
                best_score = score
                best_url   = href if href.startswith("http") else BASE_URL + href

        if best_url and best_score > 0:
            print(f"    Found product page (score={best_score}): {best_url}")
            return best_url

        print(f"    No product found for '{product_name}'")
        return None

    except Exception as e:
        print(f"    Search error: {e}")
        return None


# ── Step 2: Scrape all catalog images from product page ──────────────────────
def get_images_from_page(product_url):
    """Return list of (full_image_url, filename) — Magento cache path stripped."""
    try:
        r = requests.get(product_url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            print(f"    Product page HTTP {r.status_code}")
            return []

        soup  = BeautifulSoup(r.text, "html.parser")
        found = []
        seen  = set()

        all_srcs = [img["src"] for img in soup.find_all("img", src=True)]
        all_srcs += [a["href"] for a in soup.find_all("a", href=True)]

        for src in all_srcs:
            if "/media/catalog/product/" not in src:
                continue
            # Strip Magento cache segment to get full resolution URL
            # Before: /store/media/catalog/product/cache/1/image/HASH/a/b/file.jpg
            # After:  /store/media/catalog/product/a/b/file.jpg
            # Strip Magento cache segments — handles variable depth:
            # /cache/1/image/HASH/a/b/file.jpg  -> /a/b/file.jpg
            # /cache/1/small_image/150x/HASH/a/b/file.jpg -> /a/b/file.jpg
            clean = re.sub(
                r"(/media/catalog/product/)cache(?:/[^/]+)+/([a-z0-9]/[a-z0-9]/)",
                r"\1\2",
                src
            )
            if clean.startswith("//"):  clean = "https:" + clean
            elif clean.startswith("/"): clean = BASE_URL + clean

            if clean in seen:
                continue
            seen.add(clean)

            fname = clean.split("/")[-1].split("?")[0]
            found.append((clean, fname))
            print(f"    Found image: {fname}")

        return found

    except Exception as e:
        print(f"    Page scrape error: {e}")
        return []


# ── Step 3: Pick the correct image by count ─────────────────────────────────
def pick_image(all_images, count):
    """
    Match count against image filenames.
    Handles both:
      formula50-250.jpg  (separator before count)
      medizym800.jpg     (count directly after name letters)
    """
    if not all_images:
        return None
    if count:
        for url, fname in all_images:
            fname_noext = fname.lower().rsplit('.', 1)[0]
            if re.search(r'(?<!\d)' + re.escape(count) + r'(?!\d)', fname_noext):
                print(f"    ✓ Matched count '{count}': {fname}")
                return (url, fname)
        print(f"    WARNING: No image matched count '{count}'")
        print(f"    Available: {[f for _, f in all_images]}")
        print(f"    Falling back to first image: {all_images[0][1]}")
    return all_images[0]


# ── Step 4: Download image ────────────────────────────────────────────────────
def download_image(img_url):
    try:
        r = requests.get(img_url, headers=HEADERS, timeout=20)
        if r.status_code == 200 and len(r.content) > 1000:
            return r.content
        print(f"    HTTP {r.status_code}: {img_url}")
        return None
    except Exception as e:
        print(f"    Download error: {e}")
        return None


# ── Main: one product ─────────────────────────────────────────────────────────
def process_one(sku, upc, product_name, folder):
    product_url = search_product(product_name)
    if not product_url:
        print(f"  SKU {sku} — not found on naturally.com")
        return 0

    all_images = get_images_from_page(product_url)
    if not all_images:
        print(f"  SKU {sku} — no images on product page")
        return 0

    count  = extract_count(product_name)
    print(f"    Count from product name: {count}")

    selected = pick_image(all_images, count)
    if not selected:
        print(f"  SKU {sku} — could not select image")
        return 0

    img_url, fname = selected
    os.makedirs(folder, exist_ok=True)
    name      = f"{upc}_main"
    print(f"  Downloading {fname}...")
    img_bytes = download_image(img_url)
    if img_bytes and process_image_bytes(img_bytes, folder, name):
        return 1
    return 0


# ── Excel processing ──────────────────────────────────────────────────────────
def process_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb["Image Resize Jobs"]

    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        upc    = str(row[1].value).strip() if row[1].value else ""
        sku    = str(row[2].value).strip() if row[2].value else ""
        name   = str(row[3].value).strip() if row[3].value else ""
        url    = str(row[5].value).strip() if row[5].value else ""
        status = str(row[6].value).strip() if row[6].value else ""

        if not upc or upc == "UPC *": continue
        if status == "Done":
            print(f"Skipping {sku} — already Done"); continue
        if not url or NATURALLY_DOMAIN not in url: continue
        rows.append((sku, upc, name, row[6]))

    if not rows:
        print("No naturally.com products found."); return

    total = len(rows)
    print(f"\nFound {total} naturally.com products to process\n")

    for i, (sku, upc, name, status_cell) in enumerate(rows, 1):
        folder = os.path.join(OUTPUT_DIR, sku)
        os.makedirs(folder, exist_ok=True)
        print(f"{'='*60}")
        print(f"[{i}/{total}] SKU: {sku} | UPC: {upc} | {name}")
        print(f"{'='*60}")

        count = process_one(sku, upc, name, folder)
        status_cell.value = "Done" if count > 0 else "Failed"
        print(f"  → {count} image(s) saved\n")
        wb.save(filepath)

    print(f"\nComplete!")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Naturally Vitamins Image Downloader")
    parser.add_argument("--input",  help="Path to Excel file")
    parser.add_argument("--sku",    help="Single product SKU")
    parser.add_argument("--upc",    help="Single product UPC")
    parser.add_argument("--name",   help="Product name for search")
    parser.add_argument("--folder", help="Output folder for single product")
    args = parser.parse_args()

    if args.sku and args.folder:
        os.makedirs(args.folder, exist_ok=True)
        upc   = args.upc  if args.upc  else args.sku
        name  = args.name if args.name else args.sku
        count = process_one(args.sku, upc, name, args.folder)
        print(f"  SKU {args.sku}: {count} image(s) downloaded")
        sys.exit(0 if count > 0 else 1)

    elif args.input:
        if not os.path.exists(args.input):
            print(f"File not found: {args.input}"); sys.exit(1)
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        process_excel(args.input)

    else:
        print("Usage: naturally_downloader.py --input file.xlsx")
        print("   or: naturally_downloader.py --sku M60250 --upc 032115602503 --name 'Formula 50 250 Softgels' --folder resized_images/M60250")
        sys.exit(1)
