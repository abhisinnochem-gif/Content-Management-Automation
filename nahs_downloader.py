"""
North American Herb & Spice (northamericanherbandspice.com) Image Downloader
=============================================================================
WordPress/WooCommerce site — no Cloudflare, plain requests work fine.

Flow:
  1. Search /search-results?q={SKU} → extract product page URL
     Skip bundle/pack/set results
  2. Load product page with requests + BeautifulSoup
  3. Parse woocommerce-product-gallery__image divs → data-large_image URLs
  4. Find dominant size token (1oz, 45oz, 12oz etc.) across Front/Left/Right images
  5. Pick images matching that size:
     - Front → main
     - Right → sfp
     - Left  → left
     Skip lifestyle/award/bundle images

Usage:
    python nahs_downloader.py --sku 033 --upc 635824000136 --name "Oreganol" --folder resized_images/033
    python nahs_downloader.py --input "Product_Image_Resize_.xlsx"

Install:
    pip install requests openpyxl pillow numpy beautifulsoup4 rembg onnxruntime
"""

import os, sys, argparse, requests, re, asyncio, shutil, tempfile
from io import BytesIO
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from PIL import Image, ImageFilter, ImageEnhance
import numpy as np

OUTPUT_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resized_images")
NAHS_DOMAIN = "northamericanherbandspice.com"
BASE_URL    = "https://www.northamericanherbandspice.com"

SIZES = {
    "500x500_MasterDB":   (500,  500),
    "1000x1000_LiveSite": (1000, 1000),
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
}

# Skip images with these words in filename or alt — not product shots
SKIP_KEYWORDS = [
    "award", "lifestyle", "banner", "logo", "icon", "cert",
    "badge", "background", "bg", "texture", "pattern",
]

# Bundle/pack URL keywords to skip in search results
SKIP_URL_KEYWORDS = ["bundle", "pack", "set", "kit", "duo", "combo", "value"]


# ── Background removal ────────────────────────────────────────────────────────
def has_clean_background(img, threshold=0.50):
    arr   = np.array(img.convert("RGB"))
    white = (arr[:,:,0] > 235) & (arr[:,:,1] > 235) & (arr[:,:,2] > 235)
    return white.sum() / white.size > threshold

def remove_background(img):
    if has_clean_background(img):
        print(f"    [BG] Clean background — skipping rembg")
        return img.convert("RGBA")
    print(f"    [BG] Removing background...")
    try:
        from rembg import remove as rembg_remove, new_session
        session = new_session("u2net")
        buf = BytesIO()
        img.convert("RGB").save(buf, format="PNG")
        result = rembg_remove(buf.getvalue(), session=session, alpha_matting=False)
        return Image.open(BytesIO(result)).convert("RGBA")
    except Exception as e:
        print(f"    [BG] Failed ({e}) — using original")
        return img.convert("RGBA")


# ── Process and save image ────────────────────────────────────────────────────
def process_image_bytes(img_bytes, folder, name, img_type="main"):
    try:
        img = Image.open(BytesIO(img_bytes))
        if img.width < 100 or img.height < 100:
            print(f"    Too small ({img.width}x{img.height}) — skipping")
            return False

        print(f"    Downloaded: {img.width}x{img.height}px")
        img = remove_background(img)

        arr  = np.array(img)
        mask = arr[:,:,3] > 10
        rows = np.any(mask, axis=1)
        cols = np.any(mask, axis=0)
        if rows.any() and cols.any():
            rmin = int(np.where(rows)[0][0]);  rmax = int(np.where(rows)[0][-1])
            cmin = int(np.where(cols)[0][0]);  cmax = int(np.where(cols)[0][-1])
            ph = max(4, int((rmax - rmin) * 0.02))
            pw = max(4, int((cmax - cmin) * 0.02))
            img = img.crop((max(0,cmin-pw), max(0,rmin-ph),
                            min(img.width,cmax+pw+1), min(img.height,rmax+ph+1)))
            print(f"    [Crop] {img.width}x{img.height}")

        os.makedirs(folder, exist_ok=True)
        for label, (w, h) in SIZES.items():
            scale   = min(int(w*0.90)/img.width, int(h*0.90)/img.height)
            new_w   = int(img.width  * scale)
            new_h   = int(img.height * scale)
            resized = img.resize((new_w, new_h), Image.LANCZOS)

            rgb = resized.convert("RGB")
            rgb = rgb.filter(ImageFilter.UnsharpMask(radius=1.0, percent=120, threshold=3))
            rgb = ImageEnhance.Sharpness(rgb).enhance(1.3)
            r2,g2,b2 = rgb.split(); _,_,_,a = resized.split()
            resized = Image.merge("RGBA",(r2,g2,b2,a))

            canvas = Image.new("RGBA",(w,h),(255,255,255,255))
            canvas.paste(resized,((w-new_w)//2,(h-new_h)//2),resized)
            out = os.path.join(folder, f"{name}_{label}.jpg")
            canvas.convert("RGB").save(out, "JPEG", quality=95, dpi=(96,96))
            print(f"      → {os.path.basename(out)}")

        return True
    except Exception as e:
        print(f"    Error processing image: {e}")
        return False


# ── Step 1: Search by SKU → product URL via Crawlee ──────────────────────────
async def _crawlee_search(sku):
    """Use raw Playwright (not Crawlee crawler) to bypass 403 session blocking."""
    from playwright.async_api import async_playwright

    search_url = f"{BASE_URL}/search-results?q={sku}"
    result     = {}

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page    = await browser.new_page()
        try:
            await page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
            await page.wait_for_timeout(3000)
            html = await page.content()
            soup = BeautifulSoup(html, "html.parser")
            for a in soup.find_all("a", href=True):
                href = a["href"].strip()
                if "/shop/" not in href:
                    continue
                if not href.startswith("http"):
                    href = BASE_URL + href
                href_low = href.lower()
                if any(k in href_low for k in SKIP_URL_KEYWORDS):
                    print(f"    Skipping bundle: {href}")
                    continue
                if "result" not in href and href.count("/") >= 4:
                    result["url"] = href
                    print(f"    Found: {href}")
                    break
        except Exception as e:
            print(f"    Search error: {e}")
        finally:
            await browser.close()

    return result.get("url")

def search_product(sku):
    """Search for product URL using Crawlee to bypass Cloudflare."""
    print(f"    Searching: {BASE_URL}/search-results?q={sku}")
    return asyncio.run(_crawlee_search(sku))


# ── Step 2: Extract size token from filename ──────────────────────────────────
def extract_size_token(filename):
    """
    Extract size/variant token from filename.
    'Oreganol-1oz-Front-1900x1920.png' -> '1oz'
    'Oreganol-45oz_1920x1900_Front.png' -> '45oz'
    'Product-12oz-Left.png' -> '12oz'
    Returns token string or None.
    """
    fname = filename.lower()
    # Match patterns like 1oz, 45oz, 12oz, 0.45oz, 2oz, 500ml etc.
    m = re.search(r'(\d+\.?\d*(?:oz|ml|fl-oz|floz|g|mg|lb|ct|count|cap|tab))', fname)
    if m:
        return m.group(1)
    return None


# ── Step 3: Parse gallery images from product page ────────────────────────────
async def _crawlee_fetch_page(url):
    """Use raw Playwright to load product page."""
    from playwright.async_api import async_playwright
    result = {}
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page    = await browser.new_page()
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            await page.wait_for_timeout(3000)
            result["html"] = await page.content()
        except Exception as e:
            print(f"    Page error: {e}")
        finally:
            await browser.close()
    return result.get("html", "")

def get_images_from_page(product_url):
    """
    Parse woocommerce-product-gallery__image divs via Crawlee.
    Returns list of (full_url, img_type, size_token).
    """
    try:
        html = asyncio.run(_crawlee_fetch_page(product_url))
        if not html:
            print(f"    Failed to load product page")
            return []

        soup  = BeautifulSoup(html, "html.parser")
        divs  = soup.find_all("div", class_="woocommerce-product-gallery__image")
        print(f"    Found {len(divs)} gallery image(s)")

        all_images = []
        for div in divs:
            img = div.find("img")
            if not img:
                continue

            # Use data-large_image for full resolution
            full_url = img.get("data-large_image", "")
            if not full_url:
                # Fallback: href of parent anchor
                a = div.find("a", href=True)
                full_url = a["href"] if a else ""

            if not full_url:
                continue

            alt   = img.get("alt", "").lower()
            fname = full_url.split("/")[-1].lower()
            fname_noext = fname.rsplit(".", 1)[0]

            # Skip lifestyle/award images
            if any(k in fname or k in alt for k in SKIP_KEYWORDS):
                print(f"    Skipping: {fname}")
                continue

            # Determine image type from filename
            img_type = None
            if "front" in fname_noext:
                img_type = "main"
            elif "right" in fname_noext:
                img_type = "sfp"
            elif "left" in fname_noext:
                img_type = "left"
            else:
                continue  # skip unknown types

            size_token = extract_size_token(fname_noext)
            all_images.append((full_url, img_type, size_token))
            print(f"    Found [{img_type}] size={size_token}: {fname}")

        return all_images

    except Exception as e:
        print(f"    Page scrape error: {e}")
        return []


# ── Step 4: Pick correct size images ─────────────────────────────────────────
def pick_size_images(all_images):
    """
    Find the dominant size token across Front/Right/Left images.
    Return only images matching that size token.
    If only one size exists, return all.
    """
    if not all_images:
        return []

    # Count size tokens
    size_counts = {}
    for _, _, size in all_images:
        if size:
            size_counts[size] = size_counts.get(size, 0) + 1

    if not size_counts:
        # No size tokens found — just return all
        print(f"    No size tokens found — using all images")
        return all_images

    if len(size_counts) == 1:
        # Only one size — use all
        return all_images

    # Multiple sizes — pick the one with most images (usually the main variant)
    # If tied, pick first appearing
    dominant = max(size_counts, key=lambda s: size_counts[s])
    print(f"    Sizes found: {size_counts} → selecting '{dominant}'")

    filtered = [(url, img_type, size) for url, img_type, size in all_images if size == dominant]
    return filtered


# ── Download image via Crawlee ───────────────────────────────────────────────
async def _crawlee_download(url):
    """Use raw Playwright to download image bytes."""
    from playwright.async_api import async_playwright
    result = {}
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page    = await browser.new_page()
        try:
            response = await page.goto(url, timeout=30000)
            if response and response.status == 200:
                result["body"] = await response.body()
        except Exception as e:
            print(f"    Download error: {e}")
        finally:
            await browser.close()
    return result.get("body")

def download_image(img_url):
    try:
        # Try plain requests first (faster)
        r = requests.get(img_url, headers=HEADERS, timeout=20)
        if r.status_code == 200 and len(r.content) > 1000:
            return r.content
        # Fallback to Crawlee if blocked
        print(f"    HTTP {r.status_code} — trying Crawlee...")
        return asyncio.run(_crawlee_download(img_url))
    except Exception as e:
        print(f"    Download error: {e}")
        return asyncio.run(_crawlee_download(img_url))


# ── Main: one product ─────────────────────────────────────────────────────────
def process_one(sku, upc, product_name, folder):
    # Step 1: find product URL
    product_url = search_product(sku)
    if not product_url:
        print(f"  SKU {sku} — not found on NAHS")
        return 0

    # Step 2: get all gallery images
    all_images = get_images_from_page(product_url)
    if not all_images:
        print(f"  SKU {sku} — no images found")
        return 0

    # Step 3: pick correct size
    images = pick_size_images(all_images)
    if not images:
        print(f"  SKU {sku} — no images after size filtering")
        return 0

    # Step 4: download and process
    os.makedirs(folder, exist_ok=True)
    count = 0
    seen_types = set()
    for img_url, img_type, size in images:
        if img_type in seen_types:
            continue
        seen_types.add(img_type)
        name      = f"{upc}_{img_type}"
        print(f"  Downloading [{img_type}] {img_url.split('/')[-1]}...")
        img_bytes = download_image(img_url)
        if img_bytes and process_image_bytes(img_bytes, folder, name, img_type):
            count += 1

    return count


# ── Excel processing ──────────────────────────────────────────────────────────
def process_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb["Image Resize Jobs"]
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        upc    = str(row[1].value).strip() if row[1].value else ""
        sku    = str(row[2].value).strip() if row[2].value else ""
        name   = str(row[3].value).strip() if row[3].value else ""
        url    = str(row[5].value).strip() if row[5].value else ""
        status = str(row[6].value).strip() if row[6].value else ""
        if not upc or upc == "UPC *": continue
        if status == "Done":
            print(f"Skipping {sku} — already Done"); continue
        if not url or NAHS_DOMAIN not in url: continue
        rows.append((sku, upc, name, row[6]))

    if not rows:
        print("No NAHS products found."); return

    total = len(rows)
    print(f"\nFound {total} NAHS products to process\n")
    for i, (sku, upc, name, status_cell) in enumerate(rows, 1):
        folder = os.path.join(OUTPUT_DIR, sku)
        os.makedirs(folder, exist_ok=True)
        print(f"{'='*60}")
        print(f"[{i}/{total}] SKU: {sku} | UPC: {upc} | {name}")
        print(f"{'='*60}")
        count = process_one(sku, upc, name, folder)
        status_cell.value = "Done" if count > 0 else "Failed"
        print(f"  → {count} images saved\n")
        wb.save(filepath)
    print(f"\nComplete!")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="North American Herb & Spice Image Downloader")
    parser.add_argument("--input",  help="Path to Excel file")
    parser.add_argument("--sku",    help="Single product SKU")
    parser.add_argument("--upc",    help="Single product UPC")
    parser.add_argument("--name",   help="Product name")
    parser.add_argument("--folder", help="Output folder")
    args = parser.parse_args()

    if args.sku and args.folder:
        os.makedirs(args.folder, exist_ok=True)
        upc   = args.upc  if args.upc  else args.sku
        name  = args.name if args.name else args.sku
        count = process_one(args.sku, upc, name, args.folder)
        print(f"  SKU {args.sku}: {count} images downloaded")
        sys.exit(0 if count > 0 else 1)
    elif args.input:
        if not os.path.exists(args.input):
            print(f"File not found: {args.input}"); sys.exit(1)
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        process_excel(args.input)
    else:
        print("Usage: nahs_downloader.py --input file.xlsx")
        print("   or: nahs_downloader.py --sku 033 --upc 635824000136 --name 'Oreganol' --folder resized_images/033")
        sys.exit(1)
