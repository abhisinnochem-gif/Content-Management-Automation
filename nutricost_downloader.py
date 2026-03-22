"""
Nutricost Image Downloader — Plugin for download_webp.py
=========================================================
Exact same default flow as original download_webp.py:
  - Search by SKU then UPC, validate with BAD_SLUGS
  - Find gallery container, match images by NTC_ prefix (Priority 1)
    then SKU/UPC/name tokens (Priority 2)
  - Remove BG, crop, resize to 4 sizes, save as JPEG

Usage:
    python nutricost_downloader.py --sku NTC727101 --upc 810139575832 --name "Vitamin C" --folder resized_images/NTC727101
"""

import os, sys, argparse, re, requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse
from io import BytesIO
from openpyxl import load_workbook
from PIL import Image, ImageFilter, ImageEnhance
import numpy as np

OUTPUT_DIR       = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resized_images")
NUTRICOST_DOMAIN = "nutricost.com"
BASE_URL         = "https://nutricost.com"

SIZES = {
    "500x500_MasterDB":    (500,  500),
    "600x600":             (600,  600),
    "1000x1000_LiveSite":  (1000, 1000),
    "1200x1200_Marketing": (1200, 1200),
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
}

BAD_SLUGS = [
    "gift-card", "gift_card", "cart", "account", "login", "register",
    "search", "collections", "pages", "blogs", "about", "contact",
    "shipping", "returns", "faq", "wholesale", "rewards", "refer",
    "404", "password", "challenge",
]

SKIP = [
    "logo", "icon", "banner", "badge", "cert", "label", "clv-main",
    "vegan", "kosher", "halal", "usda", "gfco", "nongmo", "non-gmo",
    "b-corp", "qai", "vegetarian", "sustainable", "kof-k", "kof_k",
    "yesno", "open-bottle", "_copy", "claims", "lifestyle",
    "pl_sf", "sf-box", "bcorp", "b_corp", "halal_white", "usda_white",
    "cert-sustainable", "nongmoproject", "b-corp_white", "recommended",
    "related", "upsell", "cross-sell", "recently-viewed", "also-like",
]

# Nutricost product images always start with NTC_ or ntc_
VENDOR_PREFIXES = ["NTC_", "ntc_"]


def fetch(url):
    try:
        r = requests.get(url, headers=HEADERS, timeout=20, allow_redirects=True)
        if r.status_code == 200:
            return r
        print(f"  HTTP {r.status_code} for {url}")
    except Exception as e:
        print(f"  Error: {e}")
    return None


def is_valid_product_url(url):
    path = urlparse(url).path.lower()
    if not any(s in path for s in ["/products/", "/product/", "/item/", "/detail/", "/pd/"]):
        return False
    if any(bad in path for bad in BAD_SLUGS):
        return False
    return True


def get_product_url(sku, upc):
    search_terms = []
    if sku: search_terms.append(("SKU", sku))
    if upc: search_terms.append(("UPC", upc))

    for label, term in search_terms:
        search = f"{BASE_URL}/search?q={requests.utils.quote(str(term))}"
        print(f"  Searching by {label}: {search}")
        r = fetch(search)
        if not r: continue
        soup = BeautifulSoup(r.text, "html.parser")
        candidates = []
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if not href: continue
            full = href if href.startswith("http") else BASE_URL + href
            if is_valid_product_url(full):
                candidates.append(full)
        if candidates:
            term_lower = str(term).lower()
            for c in candidates:
                if term_lower in c.lower():
                    print(f"  Found exact match via {label}: {c}")
                    return c
            print(f"  Found via {label}: {candidates[0]}")
            return candidates[0]
        print(f"  No result for {label}: {term}")
    return None


def get_all_webp_urls(product_url, sku, upc, product_name):
    r = fetch(product_url)
    if not r:
        return []

    html = r.text
    soup = BeautifulSoup(r.text, "html.parser")

    tokens = []
    if sku:  tokens.append(str(sku).lower())
    if upc:  tokens.append(str(upc).lower())
    if product_name:
        for word in re.split(r"[\s/,\-]+", product_name):
            if len(word) >= 4:
                tokens.append(word.lower())
    print(f"  Match tokens: {tokens}")

    GALLERY_SELECTORS = [
        {"data-section-type": "product"},
        {"class": re.compile(r"product.*(gallery|image|media|photo)", re.I)},
        {"class": re.compile(r"(gallery|product-image|product-photo|product-media)", re.I)},
        {"id":    re.compile(r"(product-image|product-gallery|product-media)", re.I)},
    ]
    gallery_html = None
    for selector in GALLERY_SELECTORS:
        container = soup.find(attrs=selector)
        if container:
            gallery_html = str(container)
            print(f"  Found gallery container: {list(selector.keys())[0]}")
            break

    search_html = gallery_html if gallery_html else html
    if not gallery_html:
        print(f"  No gallery container found — scanning full page (will filter by token)")

    cdn_pattern = re.compile(
        r"(https?:)?//[^<>\s]+/cdn/shop/files/([^<>\s?&]+\.(?:webp|jpg|jpeg|png))",
        re.IGNORECASE
    )
    filenames = {}
    for match in cdn_pattern.finditer(search_html):
        domain_part = match.group(1) or "https:"
        fname       = match.group(2)
        full_url    = domain_part + "//" + match.group(0).split("//")[1]
        w_match     = re.search(r"width=(\d+)", full_url)
        w           = int(w_match.group(1)) if w_match else 0
        clean       = re.sub(r"_(\d+x\d*|x\d+|pico|icon|thumb|small|compact|medium|large|grande)(?=\.)", "", fname)
        if clean not in filenames or w > filenames[clean][1]:
            cdn_base = re.search(r"(https?://[^/]+/cdn/shop/files/)", full_url)
            if cdn_base:
                filenames[clean] = (cdn_base.group(1), w)

    # Nutricost uses non-Shopify CDN — fallback to all image URLs
    if not filenames:
        all_imgs = re.findall(
            r"https?://[^<>\s]+\.(?:webp|jpg|jpeg|png)(?:\?[^<>\s]*)?",
            search_html, re.IGNORECASE
        )
        for url in all_imgs:
            base  = url.split("?")[0]
            fname = base.split("/")[-1]
            clean = re.sub(r"_(\d+x\d*|x\d+|pico|icon|thumb|small|compact|medium|large|grande)(?=\.)", "", fname)
            if clean not in filenames:
                filenames[clean] = (base.rsplit("/", 1)[0] + "/", 0)

    if not filenames:
        print(f"  No images found on page")
        return []

    print(f"  Found {len(filenames)} unique filenames")

    matched  = []
    fallback = []

    for fname, (cdn_base, w) in filenames.items():
        fl = fname.lower()
        if any(k in fl for k in SKIP): continue
        if fl.endswith(".svg"): continue

        img_url = f"{cdn_base}{fname}?format=webp&width=2000"

        # Priority 1: NTC_ vendor prefix
        prefix_match = any(fname.startswith(p) for p in VENDOR_PREFIXES)
        # Priority 2: token match
        token_match  = bool(tokens and any(tok in fl for tok in tokens))

        if prefix_match or token_match:
            matched.append((img_url, fname, w))
        else:
            fallback.append((img_url, fname, w))

    matched.sort(key=lambda x: x[2], reverse=True)
    fallback.sort(key=lambda x: x[2], reverse=True)

    if matched:
        print(f"  {len(matched)} images matched (vendor prefix / SKU / UPC / name):")
        for _, fname, w in matched:
            print(f"    {fname} ({w}w)")
        return [url for url, _, _ in matched]

    if gallery_html and fallback:
        print(f"  No token matches — using {len(fallback)} gallery images")
        for _, fname, w in fallback[:5]:
            print(f"    {fname} ({w}w)")
        return [url for url, _, _ in fallback]

    print(f"  No matching images found")
    return []


def has_clean_white_background(img, threshold=0.50):
    arr   = np.array(img.convert("RGB"))
    white = (arr[:,:,0] > 240) & (arr[:,:,1] > 240) & (arr[:,:,2] > 240)
    return white.sum() / white.size > threshold


def remove_shadow(img):
    if has_clean_white_background(img):
        print(f"  [BG] Clean white background detected — skipping rembg (preserves sharpness)")
        return img.convert("RGBA")
    print(f"  [BG] Non-white background detected — running rembg...")
    try:
        from rembg import remove as rembg_remove, new_session
        session = new_session("u2net")
        buf = BytesIO()
        img.save(buf, format="PNG")
        result = rembg_remove(buf.getvalue(), session=session, alpha_matting=False)
        print(f"  [BG] Removed successfully.")
        return Image.open(BytesIO(result)).convert("RGBA")
    except ImportError:
        print("  [BG] rembg not installed.")
        return img.convert("RGBA")
    except Exception as e:
        print(f"  [BG] Failed ({e}) — using original.")
        return img.convert("RGBA")


def download_and_save(url, save_path):
    try:
        r = requests.get(url, headers=HEADERS, timeout=20)
        if r.status_code != 200:
            print(f"  HTTP {r.status_code} for {url}")
            return False

        img = Image.open(BytesIO(r.content)).convert("RGBA")
        if img.width < 100 or img.height < 100:
            print(f"  Skipping tiny image: {img.width}x{img.height}")
            return False

        print(f"  ✓ Downloaded: {img.width}x{img.height}px")
        img = remove_shadow(img)

        arr   = np.array(img.convert("RGBA"))
        alpha = arr[:, :, 3]
        mask  = alpha > 10
        rows  = np.any(mask, axis=1)
        cols  = np.any(mask, axis=0)
        if rows.any() and cols.any():
            rmin, rmax = int(np.where(rows)[0][0]),  int(np.where(rows)[0][-1])
            cmin, cmax = int(np.where(cols)[0][0]),  int(np.where(cols)[0][-1])
            pad_h = max(4, int((rmax - rmin) * 0.02))
            pad_w = max(4, int((cmax - cmin) * 0.02))
            rmin  = max(0, rmin - pad_h)
            rmax  = min(img.height - 1, rmax + pad_h)
            cmin  = max(0, cmin - pad_w)
            cmax  = min(img.width - 1,  cmax + pad_w)
            img_cropped = img.crop((cmin, rmin, cmax + 1, rmax + 1))
            print(f"  [Crop] {img.width}x{img.height} -> {img_cropped.width}x{img_cropped.height}")
        else:
            img_cropped = img
            print(f"  [Crop] No product found, using full image")

        folder = os.path.dirname(save_path)
        name   = os.path.splitext(os.path.basename(save_path))[0]
        os.makedirs(folder, exist_ok=True)

        for label, (w, h) in SIZES.items():
            target_w = int(w * 0.90)
            target_h = int(h * 0.90)
            crop_w, crop_h = img_cropped.size
            scale   = min(target_w / crop_w, target_h / crop_h)
            new_w   = int(crop_w * scale)
            new_h   = int(crop_h * scale)
            resized = img_cropped.resize((new_w, new_h), Image.LANCZOS)

            resized_rgb = resized.convert("RGB")
            resized_rgb = resized_rgb.filter(ImageFilter.UnsharpMask(radius=1.0, percent=120, threshold=3))
            resized_rgb = ImageEnhance.Sharpness(resized_rgb).enhance(1.3)
            if resized.mode == "RGBA":
                r2, g2, b2 = resized_rgb.split()
                _, _, _, a = resized.split()
                resized = Image.merge("RGBA", (r2, g2, b2, a))
            else:
                resized = resized_rgb

            canvas   = Image.new("RGBA", (w, h), (255, 255, 255, 255))
            offset_x = (w - new_w) // 2
            offset_y = (h - new_h) // 2
            canvas.paste(resized, (offset_x, offset_y),
                         resized if resized.mode == "RGBA" else None)

            out_path = os.path.join(folder, f"{name}_{label}.jpg")
            canvas.convert("RGB").save(out_path, "JPEG", quality=95, dpi=(96, 96))
            print(f"    → {os.path.basename(out_path)} ({new_w}x{new_h} on {w}x{h})")

        return True
    except Exception as e:
        print(f"  Failed to download {url}: {e}")
        return False


def process_one(sku, upc, product_name, folder):
    product_url = get_product_url(sku, upc)
    if not product_url:
        print(f"  No product page found for SKU: {sku}")
        return 0

    print(f"  Product page: {product_url}")
    image_urls = get_all_webp_urls(product_url, sku, upc, product_name)
    if not image_urls:
        print(f"  No images found for SKU: {sku}")
        return 0

    downloaded = 0
    for i, url in enumerate(image_urls, 1):
        fname       = url.split("/")[-1].split("?")[0]
        name_no_ext = os.path.splitext(fname)[0]
        save_path   = os.path.join(folder, f"{name_no_ext}.png")
        print(f"  [{i}/{len(image_urls)}] {fname}")
        if download_and_save(url, save_path):
            downloaded += 1

    print(f"\n  Downloaded {downloaded}/{len(image_urls)} images to: {folder}/")
    return downloaded


def process_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb["Image Resize Jobs"]
    total, success, failed = 0, 0, 0

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        upc    = str(row[1].value).strip() if row[1].value else ""
        sku    = str(row[2].value).strip() if row[2].value else ""
        name   = str(row[3].value).strip() if row[3].value else ""
        url    = str(row[5].value).strip() if row[5].value else ""
        status = str(row[6].value).strip() if row[6].value else ""

        if not upc or upc == "UPC *": continue
        if status == "Done":
            print(f"Skipping {sku} — already Done"); continue
        if not url or NUTRICOST_DOMAIN not in url: continue

        total += 1
        identifier = sku if sku else upc
        folder = os.path.join(OUTPUT_DIR, identifier)
        os.makedirs(folder, exist_ok=True)
        print(f"\n{'='*60}")
        print(f"SKU: {sku} | UPC: {upc} | {name}")
        print(f"{'='*60}")

        count = process_one(sku, upc, name, folder)
        row[6].value = "Done" if count > 0 else "Failed"
        if count > 0: success += 1
        else: failed += 1
        wb.save(filepath)

    print(f"\nDONE! {success} succeeded, {failed} failed out of {total} total.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Nutricost Image Downloader")
    parser.add_argument("--input",  help="Path to Excel file")
    parser.add_argument("--sku",    help="Single product SKU")
    parser.add_argument("--upc",    default="", help="Single product UPC")
    parser.add_argument("--name",   default="", help="Product name")
    parser.add_argument("--folder", help="Output folder")
    args = parser.parse_args()

    if args.sku and args.folder:
        os.makedirs(args.folder, exist_ok=True)
        count = process_one(args.sku, args.upc, args.name, args.folder)
        sys.exit(0 if count > 0 else 1)
    elif args.input:
        if not os.path.exists(args.input):
            print(f"File not found: {args.input}"); sys.exit(1)
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        process_excel(args.input)
    else:
        print("Usage: nutricost_downloader.py --sku NTC727101 --upc 810139575832 --name 'Vitamin C' --folder resized_images/NTC727101")
        sys.exit(1)
